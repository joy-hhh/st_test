import streamlit as st
import pandas as pd
import io
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="연결 재무제표 도우미", layout="wide")
st.title("📊 연결 재무제표 & 주석 대사 자동화")

# --- Session State 초기화 ---
if 'files' not in st.session_state:
    st.session_state.files = {
        "coa": None,
        "parent": None,
        "subsidiaries": [],
        "adjustment": None,
        "footnotes": []
    }
if 'results' not in st.session_state:
    st.session_state.results = {
        "consolidation_wp": None,
        "financial_position": None,
        "income_statement": None,
        "combined_footnotes": None,
        "validation_log": [],
        "caje_df": None
    }
if 'caje_generated' not in st.session_state:
    st.session_state.caje_generated = False


# --- Helper Functions ---
@st.cache_data
def to_excel(df_dict):
    """
    여러 데이터프레임을 하나의 Excel 파일 버퍼에 시트로 저장하고, 스타일을 적용합니다.
    df_dict: {'sheet_name': DataFrame} 형태의 딕셔너리
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in df_dict.items():
            if df.empty:
                continue
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            ws = writer.sheets[sheet_name]

            # 헤더 스타일 정의
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 헤더 스타일 적용
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 열 너비 및 숫자 서식 적용
            for i, column_name in enumerate(df.columns, 1): # openpyxl은 1-based index
                column_letter = get_column_letter(i)
                ws.column_dimensions[column_letter].width = 17
                
                # 원본 DataFrame에서 열 이름으로 데이터 타입 확인 (i-1 사용)
                if pd.api.types.is_numeric_dtype(df[df.columns[i-1]]):
                    # 열 전체에 서식 적용 (헤더 제외)
                    for cell in ws[column_letter][1:]:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0'
                            cell.alignment = Alignment(horizontal='right', vertical='center')

    return output.getvalue()

def log_validation(message):
    """검증 결과를 세션 상태에 기록합니다."""
    st.session_state.results["validation_log"].append(message)

# --- 사이드바 파일 업로드 ---
with st.sidebar:
    st.header("📁 파일 업로드")
    st.info("파일을 업로드하면 세션에 저장됩니다.")

    st.session_state.files["coa"] = st.file_uploader("1. CoA (계정 체계)", type="xlsx", key="coa_uploader")
    st.session_state.files["parent"] = st.file_uploader("2. 모회사 재무제표", type="xlsx", key="parent_uploader")
    st.session_state.files["subsidiaries"] = st.file_uploader("3. 자회사 재무제표 (다중 선택 가능)", type="xlsx", accept_multiple_files=True, key="subs_uploader")
    st.session_state.files["adjustment"] = st.file_uploader("4. 연결 조정 분개 (CAJE)", type="xlsx", key="adj_uploader")


# --- 탭 구성 ---
tab1, tab2, tab3 = st.tabs(["📈 연결 재무제표", "📝 주석 대사", "🔁 연결조정"])

# =================================================================================================
# --- 연결 재무제표 탭 ---
# =================================================================================================
with tab1:
    st.header("1. 연결 재무제표 생성")
    st.write("CoA, 모회사, 자회사 재무제표와 연결 조정 데이터를 통합하여 연결 재무상태표와 손익계산서를 생성합니다.")

    # --- 데이터 로드 및 처리 로직 ---
    if st.button("🚀 연결 재무제표 생성 실행", disabled=not (st.session_state.files["coa"] and st.session_state.files["parent"])):
        with st.spinner("데이터를 처리하고 있습니다... 잠시만 기다려주세요."):
            st.session_state.results["validation_log"] = [] # 로그 초기화

            # 1. 파일 로드 및 데이터 정제 (이전과 동일)
            try:
                def clean_df(df):
                    if "계정코드" in df.columns:
                        df = df.dropna(subset=["계정코드"])
                        df["계정코드"] = df["계정코드"].astype(str).str.strip().str.split('.').str[0]
                    return df
                coa_df = clean_df(pd.read_excel(st.session_state.files["coa"], dtype=str))
                parent_df = clean_df(pd.read_excel(st.session_state.files["parent"], dtype={"계정코드": str}))
                subs_dfs = [clean_df(pd.read_excel(f, dtype={"계정코드": str})) for f in st.session_state.files["subsidiaries"]]
                adj_df = clean_df(pd.read_excel(st.session_state.files["adjustment"], dtype={"계정코드": str})) if st.session_state.files["adjustment"] else pd.DataFrame(columns=['계정코드', '금액'])
            except Exception as e:
                st.error(f"파일을 읽는 중 오류가 발생했습니다: {e}")
                st.stop()

            # 2. 데이터 검증
            def check_duplicates(df, name):
                dups = df['계정코드'].value_counts().loc[lambda x: x > 1]
                if not dups.empty:
                    log_validation(f"⚠️ **[{name}]** 중복 계정코드 발견: {', '.join(dups.index)}")
            check_duplicates(parent_df, "모회사")
            for i, df in enumerate(subs_dfs):
                check_duplicates(df, f"자회사{i+1}")

            def check_missing_in_coa(df, coa_codes, name):
                missing = set(df['계정코드']) - coa_codes
                if missing:
                    log_validation(f"🚨 **[{name}]** CoA에 없는 계정코드 발견: {', '.join(sorted(list(missing)))}")
            coa_codes = set(coa_df['계정코드'])
            check_missing_in_coa(parent_df, coa_codes, "모회사")
            for i, df in enumerate(subs_dfs):
                check_missing_in_coa(df, coa_codes, f"자회사{i+1}")

            def check_balance_sheet_equation(df, column_name):
                """재무상태표 차대 검증 (자산 = 부채 + 자본)"""
                total_assets = df[df['FS_Element'] == 'A'][column_name].sum()
                total_liabilities = df[df['FS_Element'] == 'L'][column_name].sum()
                total_equity = df[df['FS_Element'] == 'E'][column_name].sum()
                difference = total_assets - (total_liabilities + total_equity)
                
                if abs(difference) > 1: # 사소한 반올림 오류는 무시
                    log_validation(f"❌ **[{column_name}]** 재무상태표 차대 불일치: {difference:,.0f}")
                else:
                    log_validation(f"✅ **[{column_name}]** 재무상태표 차대 일치")

            # 3. 데이터 병합 및 기본 계산 (이전과 동일)
            merged_df = coa_df.merge(parent_df[['계정코드', '금액']], on='계정코드', how='left').rename(columns={'금액': '모회사'})
            for i, df in enumerate(subs_dfs):
                merged_df = merged_df.merge(df[['계정코드', '금액']], on='계정코드', how='left').rename(columns={'금액': f'자회사{i+1}'})
            num_cols = merged_df.select_dtypes(include='number').columns
            merged_df[num_cols] = merged_df[num_cols].fillna(0)
            merged_df['단순합계'] = merged_df[num_cols].sum(axis=1)

            # --- 추가된 차대 검증 실행 ---
            check_balance_sheet_equation(merged_df, '모회사')
            for i in range(len(subs_dfs)):
                check_balance_sheet_equation(merged_df, f'자회사{i+1}')
            check_balance_sheet_equation(merged_df, '단순합계')
            # ------------------------------

            # 4. 연결 조정 처리
            adj_grouped = adj_df.groupby('계정코드', as_index=False)['금액'].sum()

            # 조정분개를 병합하고 결측값을 0으로 채웁니다.
            merged_df = merged_df.merge(adj_grouped.rename(columns={'금액': '연결조정'}), on='계정코드', how='left')
            merged_df['연결조정'] = merged_df['연결조정'].fillna(0)

            # FS_Element에 따라 조정 금액의 부호를 결정합니다.
            # L(부채), E(자본), R(수익)은 대변 계정이므로 조정액의 부호를 반전시킵니다.
            sign_map = {"L": -1, "E": -1, "R": -1}
            merged_df['sign'] = merged_df['FS_Element'].map(sign_map).fillna(1)
            
            # '연결조정' 열에 직접 부호를 적용하여 표시되는 값을 수정합니다.
            merged_df['연결조정'] = merged_df['연결조정'] * merged_df['sign']

            # 최종 금액을 계산합니다. (단순합계 + 부호가 적용된 조정액)
            merged_df['연결금액'] = merged_df['단순합계'] + merged_df['연결조정']
            
            # 계산에 사용된 임시 sign 열을 제거합니다.
            merged_df = merged_df.drop(columns=['sign'])

            # 4. 재무상태표 / 손익계산서 분리
            df_bs = merged_df[merged_df["FS_Element"].isin(["A", "L", "E"])].copy()
            df_pl = merged_df[merged_df["FS_Element"].isin(["R", "X"])].copy()
            con_amtcols = ['모회사'] + [f'자회사{i+1}' for i in range(len(subs_dfs))] + ['단순합계', '연결조정', '연결금액']
            code_cols = [c for c in coa_df.columns if c.startswith('L') and c.endswith('code')]
            name_cols = [c for c in coa_df.columns if c.startswith('L') and not c.endswith('code')]
            name_code_map = {row[name]: row[code] for code, name in zip(code_cols, name_cols) for _, row in coa_df.iterrows() if pd.notna(row[code]) and pd.notna(row[name])}

            # 5. ★★★ 사용자 로직 기반 재귀함수 재구현 ★★★
            def generate_fs_with_subtotals(df, name_cols, amount_cols, is_pl=False):
                df = df.copy()
                if is_pl:
                    df["sign"] = df["FS_Element"].map({"R": 1, "X": -1}).fillna(1)
                    for col in amount_cols:
                        df[col] = df[col] * df["sign"]

                def recursive_subtotal(data, current_name_cols):
                    if not current_name_cols or data.empty:
                        return data

                    current_col = current_name_cols[0]
                    remaining_cols = current_name_cols[1:]
                    
                    all_sub_dfs = []
                    for key, group in data.groupby(current_col, sort=False, dropna=False):
                        if pd.isna(key) or key == '':
                            all_sub_dfs.append(group)
                            continue

                        # 하위 레벨 먼저 재귀 호출
                        sub_df = recursive_subtotal(group, remaining_cols)
                        
                        # 현재 레벨의 합계 행 생성
                        sum_row = group.iloc[0:1].copy() # Get structure from the original group data
                        row_index = sum_row.index[0]
                        # Sum the original group data to prevent double counting subtotals from sub_df
                        sum_row.loc[row_index, amount_cols] = group[amount_cols].sum().values
                        sum_row.loc[row_index, '계정명'] = key
                        sum_row.loc[row_index, '계정코드'] = name_code_map.get(key, '')
                        # 하위 레벨 계층 정보는 공백으로 처리
                        for col in remaining_cols:
                            if col in sum_row.columns:
                                sum_row.loc[row_index, col] = ''
                        
                        # 세부내역 + 합계 순서로 합치기
                        all_sub_dfs.append(pd.concat([sub_df, sum_row], ignore_index=True))

                    return pd.concat(all_sub_dfs, ignore_index=True)

                final_df = recursive_subtotal(df, name_cols)

                if is_pl and not final_df.empty:
                    final_df[amount_cols] = final_df[amount_cols].divide(final_df['sign'], axis=0)
                    final_df = final_df.drop(columns=['sign'])
                
                return final_df

            # 6. 최종 결과 생성
            bs_final = generate_fs_with_subtotals(df_bs, name_cols, con_amtcols, is_pl=False)
            pl_final = generate_fs_with_subtotals(df_pl, name_cols, con_amtcols, is_pl=True)
            
            # 금액이 모든 열에서 0인 행 제거
            bs_final = bs_final.loc[(bs_final[con_amtcols] != 0).any(axis=1)]
            pl_final = pl_final.loc[(pl_final[con_amtcols] != 0).any(axis=1)]

            st.session_state.results['consolidation_wp'] = pd.concat([bs_final, pl_final]).reset_index(drop=True)
            st.session_state.results['con_amtcols'] = con_amtcols
            st.session_state.results['level_cols'] = code_cols + name_cols

            st.success("🎉 연결 재무제표 생성이 완료되었습니다!")

    # --- 결과 표시 ---
    if st.session_state.results["validation_log"]:
        with st.expander("🔍 데이터 검증 로그 보기"):
            for log in st.session_state.results["validation_log"]:
                st.markdown(log, unsafe_allow_html=True)

    if st.session_state.results["consolidation_wp"] is not None:
        st.subheader("📄 연결 작업底稿 (Working Paper)")
        
        con_amtcols = st.session_state.results.get('con_amtcols', [])
        display_cols = ['계정코드', '계정명'] + con_amtcols
        display_cols = [col for col in display_cols if col in st.session_state.results["consolidation_wp"].columns]
        st.dataframe(st.session_state.results["consolidation_wp"][display_cols])

        # --- 다운로드 로직 수정 ---
        level_cols_to_drop = st.session_state.results.get('level_cols', [])
        df_for_download = st.session_state.results["consolidation_wp"].drop(columns=level_cols_to_drop, errors='ignore')

        excel_data = to_excel({
            "Consolidation_WP": df_for_download
        })
        st.download_button(
            label="📥 결과 다운로드 (Excel)",
            data=excel_data,
            file_name="consolidated_fs_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    elif not (st.session_state.files["coa"] and st.session_state.files["parent"]):
        st.info("사이드바에서 CoA와 모회사 파일을 업로드한 후 '생성 실행' 버튼을 눌러주세요.")


# =================================================================================================
# --- 주석 대사 탭 ---
# =================================================================================================
with tab2:
    st.header("2. 주석 대사 (Reconciliation)")
    st.write("모회사 주석을 기준으로 자회사 주석들의 숫자 데이터를 위치 기반으로 합산하고, 연결정산표와 대사합니다.")

    footnote_parent_file = st.file_uploader("1. 모회사 주석 파일", type="xlsx")
    footnote_subs_files = st.file_uploader("2. 자회사 주석 파일 (다중 선택 가능)", type="xlsx", accept_multiple_files=True)

    if st.button("🔄 주석 대사 실행", disabled=not footnote_parent_file):
        if st.session_state.results.get("consolidation_wp") is None and footnote_subs_files:
            st.warning("대사를 위해서는 먼저 '연결 재무제표' 탭에서 '생성 실행'을 완료해야 합니다.")
            st.stop()

        with st.spinner("주석 파일을 취합하고 대사하고 있습니다..."):
            try:
                st.session_state.results['combined_footnotes'] = {}
                parent_sheets = pd.read_excel(footnote_parent_file, sheet_name=None, dtype=str)
                subs_files_data = [(Path(f.name).stem, pd.read_excel(f, sheet_name=None, dtype=str)) for f in footnote_subs_files]
                conso_map = st.session_state.results.get("consolidation_wp", pd.DataFrame()).set_index('계정코드')['연결금액'].to_dict() if st.session_state.results.get("consolidation_wp") is not None else {}

                for sheet_name, parent_df in parent_sheets.items():
                    if "주석" not in sheet_name:
                        continue

                    # ★★★ 파일 이름 기록 로직 추가 ★★★
                    all_dfs_for_sheet = []
                    parent_df_copy = parent_df.copy()
                    parent_df_copy["소스파일"] = Path(footnote_parent_file.name).stem
                    all_dfs_for_sheet.append(parent_df_copy)

                    for name, sheets in subs_files_data:
                        if sheet_name in sheets:
                            sub_df_copy = sheets[sheet_name].copy()
                            sub_df_copy["소스파일"] = name
                            all_dfs_for_sheet.append(sub_df_copy)
                    
                    should_concat = False
                    for df in all_dfs_for_sheet:
                        if len(df.columns) > 3:
                            for col_name in df.columns[2:-1]: # 마지막 소스파일 열 제외
                                if pd.to_numeric(df[col_name], errors='coerce').isna().any():
                                    should_concat = True
                                    break
                        if should_concat: break

                    if should_concat:
                        final_df = pd.concat(all_dfs_for_sheet, ignore_index=True)
                    else:
                        final_df = all_dfs_for_sheet[0].copy()
                        value_cols = final_df.columns[2:-1]
                        final_df[value_cols] = final_df[value_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
                        
                        for next_df in all_dfs_for_sheet[1:]:
                            next_value_cols = next_df.columns[2:-1]
                            next_df[next_value_cols] = next_df[next_value_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
                            final_df[value_cols] = final_df[value_cols].add(next_df[next_value_cols].values, fill_value=0)
                        
                        if footnote_subs_files:
                            final_df["소스파일"] = "combined"

                        # --- 연결조정 금액 합산 로직 ---
                        if "계정코드" in final_df.columns and st.session_state.results.get("consolidation_wp") is not None:
                            numeric_cols = final_df.select_dtypes(include='number').columns
                            if not numeric_cols.empty:
                                last_numeric_col = numeric_cols[-1]
                                
                                # 연결조정 데이터 가져오기
                                conso_wp_df = st.session_state.results["consolidation_wp"]
                                if '계정코드' in conso_wp_df.columns and '연결조정' in conso_wp_df.columns:
                                    adj_map = conso_wp_df.set_index('계정코드')['연결조정'].to_dict()
                                    
                                    # '계정코드'를 문자열로 변환하여 매핑 보장
                                    final_df['계정코드_str'] = final_df['계정코드'].astype(str).str.strip()
                                    
                                    # 조정액 적용
                                    adj_values = final_df['계정코드_str'].map(adj_map).fillna(0)
                                    final_df[last_numeric_col] += adj_values
                                    
                                    final_df = final_df.drop(columns=['계정코드_str'])

                        if "계정코드" in final_df.columns and conso_map:
                            last_numeric_col = final_df.select_dtypes(include='number').columns[-1]
                            def check_value_match(row):
                                code = str(row["계정코드"]).strip()
                                if not code: return ""
                                footnote_value = row[last_numeric_col]
                                conso_value = conso_map.get(code)
                                if conso_value is None: return "불일치 (정산표에 코드 없음)"
                                if abs(footnote_value - conso_value) < 1: return "일치"
                                else: return f"불일치 (차이: {footnote_value - conso_value:,.0f})"
                            final_df["대사결과"] = final_df.apply(check_value_match, axis=1)

                    st.session_state.results['combined_footnotes'][sheet_name] = final_df

                st.success("🎉 주석 취합 및 대사가 완료되었습니다!")

            except Exception as e:
                st.error(f"주석 처리 중 오류가 발생했습니다: {e}")

    # --- 결과 표시 ---
    if st.session_state.results.get('combined_footnotes'):
        st.subheader("📒 취합된 주석 데이터")
        for sheet_name, df in st.session_state.results['combined_footnotes'].items():
            with st.expander(f"시트: {sheet_name}", expanded=False):
                st.dataframe(df)
        footnote_excel_data = to_excel(st.session_state.results['combined_footnotes'])
        st.download_button(
            label="📥 취합된 주석 다운로드 (Excel)",
            data=footnote_excel_data,
            file_name="combined_footnotes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

# =================================================================================================
# --- 연결조정 탭 ---
# =================================================================================================
with tab3:
    st.header("3. 연결 조정 분개 생성")
    st.write("조정분개 입력 템플릿을 다운로드하여 작성한 후, 업로드하여 연결 조정 분개(CAJE)를 생성합니다.")

    # --- 조정분개 템플릿 다운로드 ---
    st.subheader("Step 1: 템플릿 다운로드")
    
    @st.cache_data
    def create_adjustment_template():
        adjustment_types = [
            ("CAJE01_채권채무제거", "Intercompany Elimination"),
            ("CAJE02_미실현이익제거", "Unrealized Profit Elimination"),
            ("CAJE03_투자자본상계", "Investment-Equity Elimination"),
            ("CAJE04_배당조정", "Dividend Adjustment"),
            ("CAJE05_감가상각조정", "Depreciation Adjustment"),
            ("CAJE06_상각비조정", "Amortization Adjustment"),
            ("CAJE07_손익조정", "Profit & Loss Adjustment"),
            ("CAJE08_회계정책조정", "Accounting Policy Adjustment"),
            ("CAJE09_지분법조정", "Equity Method Adjustment"),
            ("CAJE10_공정가치조정", "Fair Value Adjustment"),
            ("CAJE99_기타조정", "Other Adjustment"),
        ]
        columns = ["법인1", "계정1", "금액1", "법인2", "계정2", "금액2", "설명"]
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for sheet_name, _ in adjustment_types:
                df = pd.DataFrame(columns=columns)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]

                # 헤더 스타일 정의
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

                # 헤더 스타일 적용
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

                # 열 너비 조절
                for i, column_name in enumerate(df.columns, 1):
                    column_letter = get_column_letter(i)
                    ws.column_dimensions[column_letter].width = 20

        return output.getvalue()

    template_data = create_adjustment_template()
    st.download_button(
        label="📥 조정분개 입력 템플릿 다운로드 (.xlsx)",
        data=template_data,
        file_name="조정분개_입력템플릿.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    # --- 조정분개 파일 업로드 및 처리 ---
    st.subheader("Step 2: 템플릿 업로드 및 분개 생성")
    uploaded_adj_file = st.file_uploader("작성한 조정분개 템플릿을 업로드하세요.", type="xlsx", key="caje_uploader")

    if st.button("⚙️ 조정 분개 생성 실행", disabled=not (uploaded_adj_file and st.session_state.files["coa"])):
        with st.spinner("조정 분개를 생성하고 있습니다..."):
            try:
                coa_df = pd.read_excel(st.session_state.files["coa"], dtype=str)
                
                def get_fs_sign(fs_element):
                    if fs_element in ["A", "X"]: return -1
                    elif fs_element in ["L", "E", "R"]: return 1
                    else: return 1

                def build_caje_from_excel(adjustment_file, coa_df_internal):
                    fs_map = dict(zip(coa_df_internal["계정코드"], coa_df_internal["FS_Element"]))
                    xls = pd.ExcelFile(adjustment_file)
                    all_entries = []

                    for sheet in xls.sheet_names:
                        df = pd.read_excel(xls, sheet, dtype={"계정1": str,"계정2": str }).fillna("")
                        try:
                            caje_type = sheet.split("_")[0]
                        except IndexError:
                            caje_type = sheet

                        for _, row in df.iterrows():
                            설명 = row.get("설명", "")
                            # 법인1
                            code1 = str(row["계정1"]).strip()
                            if row["법인1"] and code1 and row["금액1"]:
                                fs1 = fs_map.get(code1, "")
                                try:
                                    금액1 = float(str(row["금액1"]).replace(",", ""))
                                    sign1 = get_fs_sign(fs1)
                                    all_entries.append({
                                        "조정유형": caje_type, "법인": row["법인1"], "계정코드": code1,
                                        "금액": 금액1 * sign1, "설명": 설명, "FS_Element": fs1
                                    })
                                except ValueError: pass
                            # 법인2
                            code2 = str(row["계정2"]).strip()
                            if row["법인2"] and code2 and row["금액2"]:
                                fs2 = fs_map.get(code2, "")
                                try:
                                    금액2 = float(str(row["금액2"]).replace(",", ""))
                                    sign2 = get_fs_sign(fs2)
                                    all_entries.append({
                                        "조정유형": caje_type, "법인": row["법인2"], "계정코드": code2,
                                        "금액": 금액2 * sign2, "설명": 설명, "FS_Element": fs2
                                    })
                                except ValueError: pass
                    return pd.DataFrame(all_entries)

                caje_df = build_caje_from_excel(uploaded_adj_file, coa_df)
                st.session_state.results['caje_df'] = caje_df
                st.session_state.caje_generated = True
                st.success("✅ 조정 분개 생성이 완료되었습니다!")

            except Exception as e:
                st.error(f"조정 분개 생성 중 오류가 발생했습니다: {e}")
    
    if not st.session_state.files["coa"]:
        st.warning("먼저 사이드바에서 CoA 파일을 업로드해야 합니다.")

    # --- 결과 표시 및 다운로드 ---
    if st.session_state.caje_generated:
        st.subheader("Step 3: 결과 확인 및 다운로드")
        st.dataframe(st.session_state.results['caje_df'])
        
        caje_excel_data = to_excel({"CAJE": st.session_state.results['caje_df']})
        st.download_button(
            label="📥 생성된 조정 분개(CAJE) 다운로드 (.xlsx)",
            data=caje_excel_data,
            file_name="CAJE.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        st.info("생성된 CAJE 파일은 사이드바의 '4. 연결 조정 분개 (CAJE)'에 업로드하여 연결 재무제표 생성에 사용할 수 있습니다.")
