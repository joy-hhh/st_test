import streamlit as st
import pandas as pd
import io
import os
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from contextlib import redirect_stdout

st.set_page_config(page_title="ConsolLab", page_icon="ConsolLab_logo.png", layout="wide")

# --- App Title ---
col1, col2 = st.columns([1, 5])
with col1:
    st.image("ConsolLab_logo.png", width=130)
with col2:
    st.title("ConsolLab")
    st.caption("연결 재무제표 & 주석 대사 자동화")

# --- Session State 초기화 ---
if 'files' not in st.session_state:
    st.session_state.files = {
        "coa": None,
        "parent": None,
        "subsidiaries": [],
        "adjustment": None,
        "footnotes": []
    }
if 'results' not in st.session_state:
    st.session_state.results = {
        "consolidation_wp_bs": None,
        "consolidation_wp_pl": None,
        "consolidation_wp_cf": None,
        "combined_footnotes": None,
        "validation_log": [],
        "caje_bspl_df": None,
        "caje_cf_df": None
    }
if 'caje_generated' not in st.session_state:
    st.session_state.caje_generated = False
if 'fcfs_results' not in st.session_state:
    st.session_state.fcfs_results = {
        "translated_df": None,
        "summary_df": None,
        "log": []
    }


# =================================================================================================
# --- Helper Functions ---
# =================================================================================================
@st.cache_data
def to_excel(df_dict):
    """
    여러 데이터프레임을 하나의 Excel 파일 버퍼에 시트로 저장하고, 스타일을 적용합니다.
    df_dict: {'sheet_name': DataFrame} 형태의 딕셔너리
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        for sheet_name, df in df_dict.items():
            if df is None or df.empty:
                continue
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            ws = writer.sheets[sheet_name]

            # 헤더 스타일 정의
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 헤더 스타일 적용
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 열 너비 및 숫자 서식 적용
            for i, column_name in enumerate(df.columns, 1): # openpyxl은 1-based index
                column_letter = get_column_letter(i)
                ws.column_dimensions[column_letter].width = 17
                
                if pd.api.types.is_numeric_dtype(df[df.columns[i-1]]):
                    for cell in ws[column_letter][1:]:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0'
                            cell.alignment = Alignment(horizontal='right', vertical='center')

    return output.getvalue()

def log_validation(message):
    """검증 결과를 세션 상태에 기록합니다."""
    st.session_state.results["validation_log"].append(message)

# =================================================================================================
# --- 외화FS환산용 함수 및 설정 (from fcfs_translate.py) ---
# =================================================================================================
AMOUNT_COL_CANDIDATES = ("외화금액", "FC_Amount", "Amount")
EQUITY_CARRY_COL = "이월금액"
NAME_COL_CANDIDATES = ("계정명", "Account", "Name")
RE_NEW_NAME = "이월이익잉여금(환산)"
EPS_BS = 1e-6

def _first_numeric_in_row(row):
    s = pd.to_numeric(row, errors="coerce").dropna()
    return None if s.empty else float(s.iloc[0])

def _find_col(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None

def read_rates_and_table(xlsx_path):
    all_df = pd.read_excel(xlsx_path, header=0)
    closing_rate = _first_numeric_in_row(all_df.iloc[0])
    average_rate = _first_numeric_in_row(all_df.iloc[1])
    if closing_rate is None or average_rate is None:
        raise ValueError("기말/평균환율을 2~3행(데이터 첫 2행)에서 찾지 못했습니다.")
    df = all_df.drop(index=[0, 1]).reset_index(drop=True)
    if "FS_Element" not in df.columns:
        raise ValueError("파일에 FS_Element 컬럼이 없습니다. (A/L/E/RE/R/X/PI)")
    return closing_rate, average_rate, df

def precheck_foreign_currency(df, eps=EPS_BS):
    df = df.copy()
    amount_col = _find_col(df, AMOUNT_COL_CANDIDATES)
    if amount_col is None:
        raise ValueError(f"외화금액 컬럼을 찾지 못했습니다. 후보={AMOUNT_COL_CANDIDATES}")
    df[amount_col] = pd.to_numeric(df[amount_col], errors="coerce").fillna(0.0)
    is_A  = df["FS_Element"].eq("A")
    is_L  = df["FS_Element"].eq("L")
    is_E  = df["FS_Element"].eq("E")
    is_RE = df["FS_Element"].eq("RE")
    is_R  = df["FS_Element"].eq("R")
    is_X  = df["FS_Element"].eq("X")
    a_fc  = df.loc[is_A,  amount_col].sum()
    l_fc  = df.loc[is_L,  amount_col].sum()
    e_fc  = df.loc[is_E | is_RE, amount_col].sum()
    ni_fc = df.loc[is_R, amount_col].sum() - df.loc[is_X, amount_col].sum()
    bs_gap_fc = a_fc - l_fc - e_fc
    print(f"[PRECHECK] (외화) A-L-(E+RE) = {bs_gap_fc}", "->", "OK" if abs(bs_gap_fc) < eps else "NG")
    print(f"[PRECHECK] (외화) NI_FC = {ni_fc}")
    return {"A_FC": a_fc, "L_FC": l_fc, "E_plus_RE_FC": e_fc, "NI_FC": ni_fc, "BS_GAP_FC": bs_gap_fc, "BS_OK_FC": abs(bs_gap_fc) < eps}

def translate_fcfs(df, closing_rate, average_rate, eps=EPS_BS):
    df = df.copy()
    amount_col = _find_col(df, AMOUNT_COL_CANDIDATES)
    if amount_col is None:
        raise ValueError(f"외화금액 컬럼을 찾지 못했습니다. 후보={AMOUNT_COL_CANDIDATES}")
    name_col = _find_col(df, NAME_COL_CANDIDATES)
    df[amount_col] = pd.to_numeric(df[amount_col], errors="coerce").fillna(0.0)
    if EQUITY_CARRY_COL in df.columns:
        df[EQUITY_CARRY_COL] = pd.to_numeric(df[EQUITY_CARRY_COL], errors="coerce").fillna(0.0)
    out_col = "금액"
    if out_col not in df.columns:
        df[out_col] = 0.0
    is_A  = df["FS_Element"].eq("A")
    is_L  = df["FS_Element"].eq("L")
    is_E  = df["FS_Element"].eq("E")
    is_RE = df["FS_Element"].eq("RE")
    is_R  = df["FS_Element"].eq("R")
    is_X  = df["FS_Element"].eq("X")
    is_PI = df["FS_Element"].eq("PI")
    df.loc[is_A | is_L, out_col] = df.loc[is_A | is_L, amount_col] * closing_rate
    df.loc[is_E | is_RE, out_col] = df[EQUITY_CARRY_COL] if EQUITY_CARRY_COL in df.columns else 0.0
    df.loc[is_R, out_col] = df.loc[is_R, amount_col] * average_rate
    df.loc[is_X, out_col] = df.loc[is_X, amount_col] * average_rate
    ni_krw = df.loc[is_R, out_col].sum() - df.loc[is_X, out_col].sum()
    if is_RE.any():
        re_idxs = df.index[is_RE]
        df.loc[re_idxs[0], out_col] = df.loc[re_idxs[0], out_col] + ni_krw
    else:
        new_row = {col: None for col in df.columns}
        new_row["FS_Element"] = "RE"
        new_row[out_col] = ni_krw
        new_row[amount_col] = 0.0
        if EQUITY_CARRY_COL in df.columns:
            new_row[EQUITY_CARRY_COL] = 0.0
        if name_col is not None:
            new_row[name_col] = RE_NEW_NAME
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        is_RE = df["FS_Element"].eq("RE")
    assets_sum = df.loc[df["FS_Element"].eq("A"), out_col].sum()
    liabs_sum  = df.loc[df["FS_Element"].eq("L"), out_col].sum()
    equity_sum = df.loc[df["FS_Element"].isin(["E", "RE"]), out_col].sum()
    diff = assets_sum - liabs_sum - equity_sum
    if is_PI.any():
        pi_idxs = df.index[is_PI]
        df.loc[pi_idxs, out_col] = 0.0
        df.loc[pi_idxs[0], out_col] = diff
    else:
        new_row = {col: None for col in df.columns}
        new_row["FS_Element"] = "PI"
        new_row[out_col] = diff
        new_row[amount_col] = 0.0
        if EQUITY_CARRY_COL in df.columns:
            new_row[EQUITY_CARRY_COL] = 0.0
        if name_col is not None:
            new_row[name_col] = "해외사업환산손익(PI)"
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    pi_krw = diff
    e_total_with_pi = df.loc[df["FS_Element"].isin(["E", "RE", "PI"]), out_col].sum()
    bs_gap_after = assets_sum - liabs_sum - e_total_with_pi
    print(f"[POSTCHECK] (환산 후) A-L-(E+RE+PI) = {bs_gap_after:.4f}", "  -> ", "OK" if abs(bs_gap_after)<eps else "NG")
    print(f"[POSTCHECK] A={assets_sum:.2f}, L={liabs_sum:.2f}, (E+RE+PI)={e_total_with_pi:.2f}, NI(from R&X)={ni_krw:.2f}, PI={pi_krw:.2f}")
    totals = {"A(KRW)": assets_sum, "L(KRW)": liabs_sum, "E_plus_RE_plus_PI(KRW)": e_total_with_pi, "NI(KRW from R&X)": ni_krw, "PI(KRW)": pi_krw, "A-L-(E+RE+PI) (after)": bs_gap_after}
    cols_to_check = [amount_col, out_col]
    if EQUITY_CARRY_COL in df.columns:
        cols_to_check.append(EQUITY_CARRY_COL)
    is_zero_row = (df[cols_to_check].fillna(0) == 0).all(axis=1)
    df = df[~is_zero_row].reset_index(drop=True)
    return df, totals

# --- 사이드바 파일 업로드 ---
with st.sidebar:
    st.header("📁 파일 업로드")
    st.info("파일을 업로드하면 세션에 저장됩니다.")
    st.session_state.files["coa"] = st.file_uploader("1. CoA (계정 체계)", type="xlsx", key="coa_uploader")
    st.session_state.files["parent"] = st.file_uploader("2. 모회사 재무제표 (BSPL, CF 시트 포함)", type="xlsx", key="parent_uploader")
    st.session_state.files["subsidiaries"] = st.file_uploader("3. 자회사 재무제표 (다중 선택 가능)", type="xlsx", accept_multiple_files=True, key="subs_uploader")
    st.session_state.files["adjustment"] = st.file_uploader("4. 연결 조정 분개 (BS/PL CAJE 수동 업로드용)", type="xlsx", key="adj_uploader")

# --- 탭 구성 ---
tab1, tab2, tab3, tab4 = st.tabs(["📈 연결 재무제표", "📝 주석 대사", "🔁 연결조정", "🌍 외화FS환산"])

# =================================================================================================
# --- 연결 재무제표 탭 ---
# =================================================================================================
with tab1:
    st.header("1. 연결 재무제표 생성")
    st.write("CoA, 모회사, 자회사 재무제표와 연결 조정 데이터를 통합하여 연결 재무상태표, 손익계산서, 현금흐름표를 생성합니다.")

    if st.button("🚀 연결 재무제표 생성 실행", disabled=not (st.session_state.files["coa"] and st.session_state.files["parent"])):
        with st.spinner("데이터를 처리하고 있습니다... 잠시만 기다려주세요."):
            st.session_state.results["validation_log"] = []

            try:
                def clean_df(df):
                    if "계정코드" in df.columns:
                        df = df.dropna(subset=["계정코드"])
                        df["계정코드"] = df["계정코드"].astype(str).str.strip().str.split('.').str[0]
                    return df

                def read_fs_sheets(file):
                    xls = pd.ExcelFile(file)
                    bspl_df = pd.read_excel(xls, sheet_name='BSPL', dtype={"계정코드": str}) if 'BSPL' in xls.sheet_names else pd.DataFrame()
                    cf_df = pd.read_excel(xls, sheet_name='CF', dtype={"계정코드": str}) if 'CF' in xls.sheet_names else pd.DataFrame()
                    return clean_df(bspl_df), clean_df(cf_df)

                coa_df = clean_df(pd.read_excel(st.session_state.files["coa"], sheet_name="CoA", dtype=str))
                parent_bspl_df, parent_cf_df = read_fs_sheets(st.session_state.files["parent"])
                
                subs_bspl_dfs, subs_cf_dfs = [], []
                for i, f in enumerate(st.session_state.files["subsidiaries"]):
                    bspl, cf = read_fs_sheets(f)
                    subs_bspl_dfs.append(bspl.rename(columns={'금액': f'자회사{i+1}'}))
                    subs_cf_dfs.append(cf.rename(columns={'금액': f'자회사{i+1}'}))

                caje_bspl_df = st.session_state.results.get('caje_bspl_df')
                if caje_bspl_df is None or caje_bspl_df.empty:
                    if st.session_state.files["adjustment"]:
                        caje_bspl_df = clean_df(pd.read_excel(st.session_state.files["adjustment"], dtype={"계정코드": str}))
                    else:
                        caje_bspl_df = pd.DataFrame(columns=['계정코드', '금액'])
                
            except Exception as e:
                st.error(f"파일을 읽는 중 오류가 발생했습니다: {e}")
                st.stop()

            # --- BS/PL 데이터 처리 ---
            merged_bspl_df = coa_df.merge(parent_bspl_df.rename(columns={'금액': '모회사'})[['계정코드', '모회사']], on='계정코드', how='left')
            for df in subs_bspl_dfs:
                merged_bspl_df = merged_bspl_df.merge(df[['계정코드', df.columns[-1]]], on='계정코드', how='left')
            
            bspl_val_cols = ['모회사'] + [f'자회사{i+1}' for i in range(len(subs_bspl_dfs))]
            merged_bspl_df[bspl_val_cols] = merged_bspl_df[bspl_val_cols].fillna(0)
            merged_bspl_df['단순합계'] = merged_bspl_df[bspl_val_cols].sum(axis=1)

            adj_bspl_grouped = caje_bspl_df.groupby('계정코드', as_index=False)['금액'].sum()
            merged_bspl_df = merged_bspl_df.merge(adj_bspl_grouped.rename(columns={'금액': '연결조정'}), on='계정코드', how='left')
            merged_bspl_df['연결조정'] = merged_bspl_df['연결조정'].fillna(0)
            merged_bspl_df['연결금액'] = merged_bspl_df['단순합계'] + merged_bspl_df['연결조정']

            # --- CF 데이터 처리 ---
            cf_coa_df = coa_df[coa_df['FS_Element'] == 'CF'].copy()
            merged_cf_df = cf_coa_df.merge(parent_cf_df.rename(columns={'금액': '모회사'})[['계정코드', '모회사']], on='계정코드', how='left')
            for df in subs_cf_dfs:
                merged_cf_df = merged_cf_df.merge(df[['계정코드', df.columns[-1]]], on='계정코드', how='left')

            cf_val_cols = ['모회사'] + [f'자회사{i+1}' for i in range(len(subs_cf_dfs))]
            merged_cf_df[cf_val_cols] = merged_cf_df[cf_val_cols].fillna(0)
            merged_cf_df['단순합계'] = merged_cf_df[cf_val_cols].sum(axis=1)

            caje_cf_df = st.session_state.results.get('caje_cf_df', pd.DataFrame())
            if not caje_cf_df.empty and 'CF_Map_Code' in coa_df.columns:
                caje_cf_with_map = caje_cf_df.merge(coa_df[['계정코드', 'CF_Map_Code']].dropna(), on='계정코드', how='left')
                adj_cf_grouped = caje_cf_with_map.groupby('CF_Map_Code', as_index=False)['조정금액'].sum()
                merged_cf_df = merged_cf_df.merge(adj_cf_grouped.rename(columns={'조정금액': '연결조정', 'CF_Map_Code': '계정코드'}), on='계정코드', how='left')
            else:
                merged_cf_df['연결조정'] = 0
            
            merged_cf_df['연결조정'] = merged_cf_df['연결조정'].fillna(0)
            merged_cf_df['연결금액'] = merged_cf_df['단순합계'] + merged_cf_df['연결조정']

            # --- 최종 FS 생성 ---
            df_bs = merged_bspl_df[merged_bspl_df["FS_Element"].isin(["A", "L", "E"])].copy()
            df_pl = merged_bspl_df[merged_bspl_df["FS_Element"].isin(["R", "X"])].copy()
            df_cf = merged_cf_df.copy()

            con_amtcols = ['모회사'] + [f'자회사{i+1}' for i in range(len(subs_bspl_dfs))] + ['단순합계', '연결조정', '연결금액']
            code_cols = [c for c in coa_df.columns if c.startswith('L') and c.endswith('code')]
            name_cols = [c for c in coa_df.columns if c.startswith('L') and not c.endswith('code')]
            name_code_map = {row[name]: row[code] for code, name in zip(code_cols, name_cols) for _, row in coa_df.iterrows() if pd.notna(row[code]) and pd.notna(row[name])}

            def generate_fs_with_subtotals(df, name_cols, amount_cols, is_pl=False):
                df = df.copy()
                if is_pl:
                    df["sign"] = df["FS_Element"].map({"R": 1, "X": -1}).fillna(1)
                    for col in amount_cols:
                        df[col] = df[col] * df["sign"]
                def recursive_subtotal(data, current_name_cols):
                    if not current_name_cols or data.empty: return data
                    current_col, remaining_cols = current_name_cols[0], current_name_cols[1:]
                    all_sub_dfs = []
                    for key, group in data.groupby(current_col, sort=False, dropna=False):
                        if pd.isna(key) or key == '':
                            all_sub_dfs.append(group)
                            continue
                        sub_df = recursive_subtotal(group, remaining_cols)
                        sum_row = group.iloc[0:1].copy()
                        sum_row.loc[sum_row.index[0], amount_cols] = group[amount_cols].sum().values
                        sum_row.loc[sum_row.index[0], '계정명'] = key
                        sum_row.loc[sum_row.index[0], '계정코드'] = name_code_map.get(key, '')
                        for col in remaining_cols:
                            if col in sum_row.columns: sum_row.loc[sum_row.index[0], col] = ''
                        all_sub_dfs.append(pd.concat([sub_df, sum_row], ignore_index=True))
                    return pd.concat(all_sub_dfs, ignore_index=True)
                final_df = recursive_subtotal(df, name_cols)
                if is_pl and not final_df.empty:
                    final_df[amount_cols] = final_df[amount_cols].divide(final_df['sign'], axis=0)
                    final_df = final_df.drop(columns=['sign'])
                return final_df

            bs_final = generate_fs_with_subtotals(df_bs, name_cols, con_amtcols, is_pl=False)
            pl_final = generate_fs_with_subtotals(df_pl, name_cols, con_amtcols, is_pl=True)
            cf_final = generate_fs_with_subtotals(df_cf, name_cols, con_amtcols, is_pl=False)
            
            bs_final = bs_final.loc[(bs_final[con_amtcols].abs().sum(axis=1)) > 1e-6]
            pl_final = pl_final.loc[(pl_final[con_amtcols].abs().sum(axis=1)) > 1e-6]
            cf_final = cf_final.loc[(cf_final[con_amtcols].abs().sum(axis=1)) > 1e-6]

            st.session_state.results['consolidation_wp_bs'] = bs_final
            st.session_state.results['consolidation_wp_pl'] = pl_final
            st.session_state.results['consolidation_wp_cf'] = cf_final
            st.session_state.results['con_amtcols'] = con_amtcols

            st.success("🎉 연결 재무제표 생성이 완료되었습니다!")

    # --- 결과 표시 ---
    if st.session_state.results.get("consolidation_wp_bs") is not None:
        st.subheader("📄 연결 재무상태표")
        st.dataframe(st.session_state.results['consolidation_wp_bs'])
        st.subheader("📄 연결 손익계산서")
        st.dataframe(st.session_state.results['consolidation_wp_pl'])
        st.subheader("📄 연결 현금흐름표")
        st.dataframe(st.session_state.results['consolidation_wp_cf'])

        excel_data = to_excel({
            "Consol_BS": st.session_state.results['consolidation_wp_bs'],
            "Consol_PL": st.session_state.results['consolidation_wp_pl'],
            "Consol_CF": st.session_state.results['consolidation_wp_cf']
        })
        st.download_button(
            label="📥 전체 결과 다운로드 (Excel)",
            data=excel_data,
            file_name="consolidated_fs_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    elif not (st.session_state.files["coa"] and st.session_state.files["parent"]):
        st.info("사이드바에서 CoA와 모회사 파일을 업로드한 후 '생성 실행' 버튼을 눌러주세요.")

# =================================================================================================
# --- 주석 대사 탭 ---
# =================================================================================================
with tab2:
    st.header("2. 주석 대사 (Reconciliation)")
    st.write("모회사 주석을 기준으로 자회사 주석들의 숫자 데이터를 위치 기반으로 합산하고, 연결정산표와 대사합니다.")
    footnote_parent_file = st.file_uploader("1. 모회사 주석 파일", type="xlsx")
    footnote_subs_files = st.file_uploader("2. 자회사 주석 파일 (다중 선택 가능)", type="xlsx", accept_multiple_files=True)
    if st.button("🔄 주석 대사 실행", disabled=not footnote_parent_file):
        if st.session_state.results.get("consolidation_wp_bs") is None and footnote_subs_files:
            st.warning("대사를 위해서는 먼저 '연결 재무제표' 탭에서 '생성 실행'을 완료해야 합니다.")
            st.stop()
        with st.spinner("주석 파일을 취합하고 대사하고 있습니다..."):
            try:
                st.session_state.results['combined_footnotes'] = {}
                parent_sheets = pd.read_excel(footnote_parent_file, sheet_name=None, dtype=str)
                subs_files_data = [(Path(f.name).stem, pd.read_excel(f, sheet_name=None, dtype=str)) for f in footnote_subs_files]
                
                conso_wp_df = pd.concat([
                    st.session_state.results.get('consolidation_wp_bs', pd.DataFrame()),
                    st.session_state.results.get('consolidation_wp_pl', pd.DataFrame()),
                    st.session_state.results.get('consolidation_wp_cf', pd.DataFrame())
                ])
                conso_map = conso_wp_df.set_index('계정코드')['연결금액'].to_dict() if not conso_wp_df.empty else {}

                for sheet_name, parent_df in parent_sheets.items():
                    if "주석" not in sheet_name: continue
                    all_dfs_for_sheet = []
                    parent_df_copy = parent_df.copy()
                    parent_df_copy["소스파일"] = Path(footnote_parent_file.name).stem
                    all_dfs_for_sheet.append(parent_df_copy)
                    for name, sheets in subs_files_data:
                        if sheet_name in sheets:
                            sub_df_copy = sheets[sheet_name].copy()
                            sub_df_copy["소스파일"] = name
                            all_dfs_for_sheet.append(sub_df_copy)
                    should_concat = any(pd.to_numeric(df[col], errors='coerce').isna().any() for df in all_dfs_for_sheet for col in df.columns[2:-1] if len(df.columns) > 3)
                    if should_concat:
                        final_df = pd.concat(all_dfs_for_sheet, ignore_index=True)
                    else:
                        final_df = all_dfs_for_sheet[0].copy()
                        value_cols = final_df.columns[2:-1]
                        final_df[value_cols] = final_df[value_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
                        for next_df in all_dfs_for_sheet[1:]:
                            next_value_cols = next_df.columns[2:-1]
                            next_df[next_value_cols] = next_df[next_value_cols].apply(pd.to_numeric, errors='coerce').fillna(0)
                            final_df[value_cols] = final_df[value_cols].add(next_df[next_value_cols].values, fill_value=0)
                        if footnote_subs_files: final_df["소스파일"] = "combined"
                        if "계정코드" in final_df.columns and not conso_wp_df.empty:
                            numeric_cols = final_df.select_dtypes(include='number').columns
                            if not numeric_cols.empty:
                                last_numeric_col = numeric_cols[-1]
                                if '연결조정' in conso_wp_df.columns:
                                    adj_map = conso_wp_df.set_index('계정코드')['연결조정'].to_dict()
                                    final_df['계정코드_str'] = final_df['계정코드'].astype(str).str.strip()
                                    adj_values = final_df['계정코드_str'].map(adj_map).fillna(0)
                                    final_df[last_numeric_col] += adj_values
                                    final_df = final_df.drop(columns=['계정코드_str'])
                        if "계정코드" in final_df.columns and conso_map:
                            last_numeric_col = final_df.select_dtypes(include='number').columns[-1]
                            def check_value_match(row):
                                code = str(row["계정코드"]).strip()
                                if not code: return ""
                                footnote_value = row[last_numeric_col]
                                conso_value = conso_map.get(code)
                                if conso_value is None: return "불일치 (정산표에 코드 없음)"
                                if abs(footnote_value - conso_value) < 1: return "일치"
                                else: return f"불일치 (차이: {footnote_value - conso_value:,.0f})"
                            final_df["대사결과"] = final_df.apply(check_value_match, axis=1)
                    st.session_state.results['combined_footnotes'][sheet_name] = final_df
                st.success("🎉 주석 취합 및 대사가 완료되었습니다!")
            except Exception as e:
                st.error(f"주석 처리 중 오류가 발생했습니다: {e}")

    if st.session_state.results.get('combined_footnotes'):
        st.subheader("📒 취합된 주석 데이터")
        for sheet_name, df in st.session_state.results['combined_footnotes'].items():
            with st.expander(f"시트: {sheet_name}", expanded=False):
                st.dataframe(df)
        footnote_excel_data = to_excel(st.session_state.results['combined_footnotes'])
        st.download_button(label="📥 취합된 주석 다운로드 (Excel)", data=footnote_excel_data, file_name="combined_footnotes.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# =================================================================================================
# --- 연결조정 탭 ---
# =================================================================================================
with tab3:
    st.header("3. 연결 조정 분개 생성")
    st.write("조정분개 입력 템플릿을 다운로드하여 작성한 후, 업로드하여 BS/IS 및 현금흐름표(CF) 목적의 연결 조정 분개(CAJE)를 생성합니다.")

    st.subheader("Step 1: 템플릿 다운로드")
    @st.cache_data
    def create_adjustment_template():
         # 'Info' 시트에 들어갈 데이터프레임을 먼저 정의
        info_data = {
            '당기세율': ['20%', '18%', '16%'],
            '전기세율': ['22%', '20%', '18%'],
            '당기지분율': ['100%', '60%', '80%'],
            '전기지분율': ['100%', '60%', '80%']
        }
        info_index_labels = ['모회사', '자회사A', '자회사B']
        info_df = pd.DataFrame(info_data, index=info_index_labels)
        info_df.index.name = 'Info'
        
        adjustment_types = [("CAJE00_투자자본상계", "Investment-Equity Elimination"), ("CAJE01_채권채무제거", "Intercompany Elimination"), ("CAJE02_제품미실현이익제거", "Unrealized Profit Elimination"), ("CAJE03_상각자산미실현이익제거", "Depreciable Assets Unrealized Profit Elimination"), ("CAJE04_배당조정", "Dividend Adjustment"), ("CAJE05_기타손익조정", "Other P&L Adjustments"), ("CAJE99_기타조정", "Other Adjustments")]
        columns = ["법인", "계정코드", "계정명", "당기전기", "금액", "설명"]
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # 'Info' 데이터프레임을 'Info'라는 시트 이름으로 엑셀에 먼저
            info_df.to_excel(writer, sheet_name="Info", index=True)
            for sheet_name, _ in adjustment_types:
                if sheet_name == "CAJE00_투자자본상계":
                    example_data = [
                        {"법인": "모회사", "계정코드": "19200", "계정명": "종속기업투자", "당기전기": "취득일", "금액": 5000000, "설명": "자회사A 투자금액 제거"},
                        {"법인": "자회사A", "계정코드": "33100", "계정명": "자본금", "당기전기": "취득일", "금액": 3000000, "설명": "자회사A 자본금 제거"},
                        {"법인": "자회사A", "계정코드": "37500", "계정명": "이익잉여금", "당기전기": "취득일", "금액": 1000000, "설명": "자회사A 이익잉여금(취득시점) 제거"},
                        {"법인": "자회사A", "계정코드": "101000", "계정명": "영업권", "당기전기": "취득일", "금액": 1000000, "설명": "자회사A 영업권 계상"},
                    ]
                    df = pd.DataFrame(example_data)
                elif sheet_name == "CAJE01_채권채무제거":
                    example_data = [
                        {"법인": "모회사", "계정코드": "10800", "계정명": "매출채권", "당기전기": "전기", "금액": 20000000, "설명": "자회사A에 대한 매출채권 제거"},
                        {"법인": "자회사A", "계정코드": "25100", "계정명": "매입채무", "당기전기": "전기", "금액": 20000000, "설명": "모회사에 대한 매입채무 제거"},
                        {"법인": "모회사", "계정코드": "10800", "계정명": "매출채권", "당기전기": "당기", "금액": 10000000, "설명": "자회사A에 대한 매출채권 제거"},
                        {"법인": "자회사A", "계정코드": "25100", "계정명": "매입채무", "당기전기": "당기", "금액": 10000000, "설명": "모회사에 대한 매입채무 제거"},
                    ]
                    df = pd.DataFrame(example_data)
                elif sheet_name == "CAJE02_제품미실현이익제거":
                    example_data = [
                        {"법인": "자회사A", "계정코드": "45500", "계정명": "매출원가", "당기전기": "전기", "금액": 3000000, "설명": "전기 미실현이익"},
                        {"법인": "자회사A", "계정코드": "37500", "계정명": "이익잉여금", "당기전기": "전기", "금액": 3000000, "설명": "전기 미실현이익(이익잉여금)"},
                        {"법인": "모회사", "계정코드": "40200", "계정명": "매출", "당기전기": "당기", "금액": 10000000, "설명": "당기 판매분 매출"},
                        {"법인": "모회사", "계정코드": "45500", "계정명": "매출원가", "당기전기": "당기", "금액": 6000000, "설명": "당기 판매분 매출원가"},
                        {"법인": "모회사", "계정코드": "15200", "계정명": "제품(재고자산)", "당기전기": "당기", "금액": 4000000, "설명": "모회사가 판매한 재고 미실현이익 제거(재고감소)"},
                    ]
                    df = pd.DataFrame(example_data)
                elif sheet_name == "CAJE03_상각자산미실현이익제거":
                    example_data = [
                        {"법인": "자회사A", "계정코드": "20600", "계정명": "기계장치", "당기전기": "전기", "금액": 5000000, "설명": "자회사A에서 모회사에 처분"},
                        {"법인": "자회사A", "계정코드": "37500", "계정명": "이익잉여금", "당기전기": "전기", "금액": 5000000, "설명": "자회사A 계상 유형자산처분이익"},
                        {"법인": "자회사A", "계정코드": "37500", "계정명": "이익잉여금", "당기전기": "전기", "금액": 1000000, "설명": "모회사 계상 감가상각비 증분 제거"},
                        {"법인": "자회사A", "계정코드": "20700", "계정명": "기계장치감가상각누계액", "당기전기": "전기", "금액": 1000000, "설명": "모회사 감가상각누계액"},
                        {"법인": "자회사A", "계정코드": "81800", "계정명": "감가상각비", "당기전기": "당기", "금액": 1000000, "설명": "모회사 계상 감가상각비 증분 제거"},
                        {"법인": "자회사A", "계정코드": "20700", "계정명": "기계장치감가상각누계액", "당기전기": "당기", "금액": 1000000, "설명": "모회사 감가상각누계액"},
                    ]
                    df = pd.DataFrame(example_data)
                elif sheet_name == "CAJE04_배당조정":
                    example_data = [
                        {"법인": "모회사", "계정코드": "90300", "계정명": "배당금수익", "당기전기": "당기", "금액": 2000000, "설명": "자회사A로부터 받은 배당금수익 제거"},
                        {"법인": "자회사A", "계정코드": "37500", "계정명": "이익잉여금", "당기전기": "당기", "금액": -2000000, "설명": "모회사에 지급한 배당금 효과 제거"},
                    ]
                    df = pd.DataFrame(example_data)
                else:
                    df = pd.DataFrame(columns=columns)
                df = df.reindex(columns=columns)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                for i, column_name in enumerate(df.columns, 1):
                    ws.column_dimensions[get_column_letter(i)].width = 20
        return output.getvalue()
    template_data = create_adjustment_template()
    st.download_button(label="📥 조정분개 입력 템플릿 다운로드 (.xlsx)", data=template_data, file_name="조정분개_입력템플릿.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    st.subheader("Step 2: 템플릿 업로드 및 분개 생성")
    uploaded_adj_file = st.file_uploader("작성한 조정분개 템플릿을 업로드하세요.", type="xlsx", key="caje_uploader")

    def build_caje_from_template(adjustment_file, coa_df_internal):
        fs_map = dict(zip(coa_df_internal["계정코드"], coa_df_internal["FS_Element"]))
        def get_bspl_sign(fs_element):
            return -1 if fs_element in ["A", "X"] else 1
        
        ni_code = None
        try:
            r_rows = coa_df_internal[coa_df_internal['FS_Element'] == 'R']
            if not r_rows.empty:
                ni_code = r_rows.iloc[0].get('L1_code')
        except (IndexError, KeyError):
            ni_code = None

        xls = pd.ExcelFile(adjustment_file)
        all_bspl_entries, all_cf_entries = [], []

        for sheet_name in xls.sheet_names:
            if not sheet_name.upper().startswith("CAJE"): continue
            caje_type = sheet_name.split("_")[0].upper()
            df = pd.read_excel(xls, sheet_name, dtype={"계정코드": str}).fillna("")

            # --- A. BS/PL Adjustment Logic ---
            df_for_bspl = df.copy()
            if caje_type in ['CAJE01', 'CAJE04', 'CAJE05', 'CAJE99']:
                df_for_bspl = df[df['당기전기'] == '당기']
            
            for _, row in df_for_bspl.iterrows():
                acc_code = str(row.get("계정코드", "")).strip()
                if not acc_code: continue
                fs_element = fs_map.get(acc_code, "")
                amount = pd.to_numeric(row.get("금액"), errors='coerce')
                if pd.isna(amount) or amount == 0: continue
                
                final_amount = amount * get_bspl_sign(fs_element)
                all_bspl_entries.append({
                    "조정유형": caje_type, "법인": row.get("법인"), "계정코드": acc_code,
                    "금액": final_amount, "설명": row.get("설명"), "FS_Element": fs_element
                })

            # --- B. CF Adjustment Logic ---
            if caje_type == "CAJE02":
                if ni_code is None:
                    st.warning(f"[{sheet_name}] CF조정 건너뜀: 당기순이익 계정 코드를 CoA에서 찾을 수 없습니다.")
                    continue
                df_with_fs = df.merge(coa_df_internal[['계정코드', 'FS_Element']], on='계정코드', how='left')
                pl_rows = df_with_fs[df_with_fs['FS_Element'].isin(['R', 'X'])].copy()
                bs_rows = df_with_fs[df_with_fs['FS_Element'].isin(['A', 'L', 'E'])].copy()
                if pl_rows.empty or bs_rows.empty:
                    st.warning(f"[{sheet_name}] CF조정 건너뜀: 시트에서 손익(R/X) 또는 재무상태(A/L/E) 계정을 찾을 수 없습니다.")
                    continue

                pl_pivot = pl_rows.pivot_table(index=['계정코드', 'FS_Element'], columns='당기전기', values='금액', aggfunc='sum').fillna(0)
                if '당기' not in pl_pivot.columns: pl_pivot['당기'] = 0
                if '전기' not in pl_pivot.columns: pl_pivot['전기'] = 0
                pl_pivot['change'] = pl_pivot['당기'] + pl_pivot['전기']
                pl_pivot['impact'] = pl_pivot.apply(lambda r: r['change'] if r.name[1] == 'X' else -r['change'], axis=1)
                total_pl_impact = pl_pivot['impact'].sum()
                
                inventory_acc_code = bs_rows.iloc[0]['계정코드']
                corp_name = df.iloc[0]['법인']

                # Line 1: NI Entry (+)
                all_cf_entries.append({
                    "조정유형": caje_type, "법인": corp_name, "계정코드": ni_code,
                    "조정금액": total_pl_impact, "설명": "[비현금손익] 미실현이익(NI)", "원계정_FS_Element": "R"
                })
                # Line 2: Inventory Entry (-)
                all_cf_entries.append({
                    "조정유형": caje_type, "법인": corp_name, "계정코드": inventory_acc_code,
                    "조정금액": -total_pl_impact, "설명": "[비현금손익] 미실현이익(재고)", "원계정_FS_Element": "A"
                })
            elif caje_type == "CAJE03":
                if ni_code is None:
                    st.warning(f"[{sheet_name}] CF조정 건너뜀: 당기순이익 계정 코드를 CoA에서 찾을 수 없습니다.")
                    continue
                df_with_fs = df.merge(coa_df_internal[['계정코드', 'FS_Element']], on='계정코드', how='left')
                pl_rows = df_with_fs[df_with_fs['FS_Element'].isin(['X', 'R'])].copy()
                bs_rows = df_with_fs[df_with_fs['FS_Element'].isin(['A', 'L', 'E'])].copy()
                if pl_rows.empty or bs_rows.empty:
                    st.warning(f"[{sheet_name}] CF조정 건너뜀: 시트에서 손익(R/X) 또는 재무상태(A/L/E) 계정을 찾을 수 없습니다.")
                    continue

                pl_pivot = pl_rows.pivot_table(index=['계정코드', 'FS_Element'], columns='당기전기', values='금액', aggfunc='sum').fillna(0)
                if '당기' not in pl_pivot.columns: pl_pivot['당기'] = 0
                if '전기' not in pl_pivot.columns: pl_pivot['전기'] = 0
                pl_pivot['change'] = pl_pivot['당기'] + pl_pivot['전기']
                pl_pivot['impact'] = pl_pivot.apply(lambda r: r['change'] if r.name[1] == 'X' else -r['change'], axis=1)
                total_pl_impact = pl_pivot['impact'].sum()
                
                pl_acc_code = pl_rows.iloc[0]['계정코드']
                corp_name = df.iloc[0]['법인']

                # Line 1: NI Entry (+)
                all_cf_entries.append({
                    "조정유형": caje_type, "법인": corp_name, "계정코드": ni_code,
                    "조정금액": total_pl_impact, "설명": "[비현금손익] 미실현이익(NI)", "원계정_FS_Element": "R"
                })
                # Line 2: Inventory Entry (-)
                all_cf_entries.append({
                    "조정유형": caje_type, "법인": corp_name, "계정코드": pl_acc_code,
                    "조정금액": -total_pl_impact, "설명": "[비현금손익] 미실현이익(손익)", "원계정_FS_Element": "X"
                })
            else:
                grouped = df.groupby(["법인", "계정코드", "설명"])
                for (corp, acc_code, desc), group in grouped:
                    pivot_df = group.pivot_table(columns="당기전기", values="금액", aggfunc='sum')
                    current_amt = pivot_df["당기"].item() if "당기" in pivot_df.columns else 0
                    prior_amt = pivot_df["전기"].item() if "전기" in pivot_df.columns else 0
                    fs_element = fs_map.get(acc_code, "")
                    
                    cf_adj_amt, cf_desc = 0, desc
                    if caje_type == "CAJE00":
                        change_amt = current_amt - prior_amt
                        if fs_element == 'L':
                            cf_adj_amt = change_amt
                        else: # For 'A' and others
                            cf_adj_amt = -change_amt
                        cf_desc = f"[운전자본] {desc}"
                    elif caje_type == "CAJE04":
                        cf_adj_amt, cf_desc = current_amt, f"[손익/재무활동] {desc}"
                    elif caje_type == "CAJE05":
                        cf_adj_amt, cf_desc = current_amt, f"[비현금손익] {desc}"

                    if abs(cf_adj_amt) > 1e-6:
                        all_cf_entries.append({
                            "조정유형": caje_type, "법인": corp, "계정코드": acc_code,
                            "조정금액": cf_adj_amt, "설명": cf_desc, "원계정_FS_Element": fs_element
                        })

        bspl_cols = ['조정유형', '법인', '계정코드', '금액', '설명', 'FS_Element']
        cf_cols = ['조정유형', '법인', '계정코드', '조정금액', '설명', '원계정_FS_Element']
        
        caje_bspl_df = pd.DataFrame(all_bspl_entries, columns=bspl_cols) if all_bspl_entries else pd.DataFrame(columns=bspl_cols)
        caje_cf_df = pd.DataFrame(all_cf_entries, columns=cf_cols) if all_cf_entries else pd.DataFrame(columns=cf_cols)

        return caje_bspl_df, caje_cf_df

    if st.button("⚙️ 조정 분개 생성 실행", disabled=not (uploaded_adj_file and st.session_state.files["coa"])):
        with st.spinner("조정 분개를 생성하고 있습니다..."):
            try:
                coa_df = pd.read_excel(st.session_state.files["coa"], sheet_name="CoA", dtype=str)
                caje_bspl_df, caje_cf_df = build_caje_from_template(uploaded_adj_file, coa_df)
                st.session_state.results['caje_bspl_df'] = caje_bspl_df
                st.session_state.results['caje_cf_df'] = caje_cf_df
                st.session_state.caje_generated = True
                st.success("✅ 조정 분개 생성이 완료되었습니다!")
            except Exception as e:
                st.error(f"조정 분개 생성 중 오류가 발생했습니다: {e}")
                st.exception(e)
    if not st.session_state.files["coa"]:
        st.warning("먼저 사이드바에서 CoA 파일을 업로드해야 합니다.")

    if st.session_state.caje_generated:
        st.subheader("Step 3: 결과 확인 및 다운로드")
        st.markdown("#### 📄 재무상태표/손익계산서 조정 분개 (BS/PL CAJE)")
        st.dataframe(st.session_state.results.get('caje_bspl_df'))
        st.markdown("#### 🌊 현금흐름표 조정 분개 (CF CAJE)")
        st.dataframe(st.session_state.results.get('caje_cf_df'))
        caje_excel_data = to_excel({"CAJE_BSPL": st.session_state.results.get('caje_bspl_df', pd.DataFrame()), "CAJE_CF": st.session_state.results.get('caje_cf_df', pd.DataFrame())})
        st.download_button(label="📥 생성된 조정 분개(CAJE) 다운로드 (.xlsx)", data=caje_excel_data, file_name="CAJE_generated.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        st.info("생성된 BS/PL CAJE 데이터는 '연결 재무제표' 탭의 '연결 조정' 데이터로 사용할 수 있습니다.")

# =================================================================================================
# --- 외화FS환산 탭 ---
# =================================================================================================
with tab4:
    st.header("4. 외화 재무제표 환산")
    st.write("외화로 작성된 재무제표(FS) 파일을 업로드하면, 지정된 환율에 따라 원화로 환산하고 결과를 표시합니다.")
    st.subheader("Step 1: 파일 업로드")
    st.info("환산할 FS파일을 업로드해주세요. 파일의 첫 두 데이터 행에는 기말환율과 평균환율이 포함되어야 합니다.")
    fcfs_file = st.file_uploader("외화 FS 파일", type="xlsx", key="fcfs_uploader")
    st.subheader("Step 2: 환산 실행")
    if st.button("⚙️ 외화FS 환산 실행", disabled=not fcfs_file):
        with st.spinner("외화 재무제표를 환산하는 중입니다..."):
            try:
                log_stream = io.StringIO()
                with redirect_stdout(log_stream):
                    closing_rate, average_rate, df = read_rates_and_table(fcfs_file)
                    pre_summary = precheck_foreign_currency(df)
                    translated_df, totals_summary = translate_fcfs(df, closing_rate, average_rate)
                log_contents = log_stream.getvalue()
                st.session_state.fcfs_results['log'] = log_contents.strip().split('\n')
                rates_summary_df = pd.DataFrame({"항목": ["기말환율", "평균환율"], "값": [closing_rate, average_rate]})
                pre_summary_df = pd.DataFrame({"항목": list(pre_summary.keys()), "값": list(pre_summary.values())})
                totals_summary_df = pd.DataFrame({"항목": list(totals_summary.keys()), "값": list(totals_summary.values())})
                summary_df = pd.concat([rates_summary_df, pre_summary_df, totals_summary_df], ignore_index=True)
                summary_df['값'] = summary_df['값'].astype(str)
                st.session_state.fcfs_results['translated_df'] = translated_df
                st.session_state.fcfs_results['summary_df'] = summary_df
                st.success("🎉 외화 재무제표 환산이 완료되었습니다!")
            except Exception as e:
                st.error(f"환산 중 오류가 발생했습니다: {e}")
                st.exception(e)
    st.subheader("Step 3: 결과 확인 및 다운로드")
    if st.session_state.fcfs_results.get("log"):
        with st.expander("🔍 처리 로그 보기"):
            st.code('\n'.join(st.session_state.fcfs_results["log"]))
    if st.session_state.fcfs_results.get("translated_df") is not None:
        st.markdown("#### 📄 환산된 재무제표")
        st.dataframe(st.session_state.fcfs_results["translated_df"])
        st.markdown("#### 📊 환산 요약")
        st.dataframe(st.session_state.fcfs_results["summary_df"])
        excel_data = to_excel({"translated": st.session_state.fcfs_results["translated_df"], "summary": st.session_state.fcfs_results["summary_df"]})
        st.download_button(label="📥 환산 결과 다운로드 (Excel)", data=excel_data, file_name="FCFS_translated.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")