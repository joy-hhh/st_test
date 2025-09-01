import streamlit as st
import pandas as pd
import io
from pathlib import Path
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from contextlib import redirect_stdout

st.set_page_config(
    page_title="ConsolLab", page_icon="ConsolLab_logo.png", layout="wide"
)

# --- App Title ---
col1, col2 = st.columns([1, 5])
with col1:
    st.image("ConsolLab_logo.png", width=130)
with col2:
    st.title("ConsolLab")
    st.caption("연결봇 ConsolLab : 연결 재무제표 자동화 생성기")

# --- Session State 초기화 ---
if "files" not in st.session_state:
    st.session_state.files = {
        "coa": None,
        "parent": None,
        "subsidiaries": [],
        "adjustment": None,
        "footnotes": [],
    }
if "results" not in st.session_state:
    st.session_state.results = {
        "consolidation_wp_bs": None,
        "consolidation_wp_pl": None,
        "consolidation_wp_cf": None,
        "combined_footnotes": None,
        "validation_log": [],
        "caje_bspl_df": None,
        "caje_cf_df": None,
    }
if "caje_generated" not in st.session_state:
    st.session_state.caje_generated = False
if "fcfs_results" not in st.session_state:
    st.session_state.fcfs_results = {
        "translated_df": None,
        "summary_df": None,
        "log": [],
    }


# =================================================================================================
# --- Helper Functions ---
# =================================================================================================
@st.cache_data
def to_excel(df_dict):
    """
    여러 데이터프레임을 하나의 Excel 파일 버퍼에 시트로 저장하고, 스타일을 적용합니다.
    df_dict: {'sheet_name': DataFrame} 형태의 딕셔너리
    """
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in df_dict.items():
            if df is None:  # df.empty 조건 제거하여 빈 시트도 생성
                continue

            # 소계 행 정보 추출 후, 'is_subtotal' 열은 엑셀에서 제외
            is_subtotal_col = df["is_subtotal"] if "is_subtotal" in df.columns else None
            df_to_write = (
                df.drop(columns=["is_subtotal"]) if is_subtotal_col is not None else df
            )
            df_to_write.to_excel(writer, sheet_name=sheet_name, index=False)

            ws = writer.sheets[sheet_name]

            # 헤더 스타일 정의
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            header_alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

            # 헤더 스타일 적용
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # 소계 행에 볼드체 적용
            if is_subtotal_col is not None:
                bold_font = Font(bold=True)
                for row_idx, is_sub in enumerate(is_subtotal_col):
                    if is_sub:
                        # openpyxl 행은 1-based, 헤더가 있으므로 +2
                        for cell in ws[row_idx + 2]:
                            cell.font = bold_font

            # 열 너비 및 숫자 서식 적용
            for i, column_name in enumerate(
                df_to_write.columns, 1
            ):  # openpyxl은 1-based index
                column_letter = get_column_letter(i)
                ws.column_dimensions[column_letter].width = 17

                if pd.api.types.is_numeric_dtype(
                    df_to_write[df_to_write.columns[i - 1]]
                ):
                    # 시트 이름에 따라 다른 숫자 서식 적용
                    number_format = "0.000" if sheet_name == "Info" else "#,##0"
                    for cell in ws[column_letter][1:]:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = number_format
                            cell.alignment = Alignment(
                                horizontal="right", vertical="center"
                            )

    return output.getvalue()

def parse_percent(s):
    """
    다양한 형태의 퍼센트 값을 소수점 형태로 변환합니다.
    - '60%': 0.6
    - 60: 0.6 (1보다 크므로 퍼센트로 간주)
    - 0.6: 0.6 (1보다 작거나 같으므로 소수점으로 간주)
    """
    # 1. 입력값이 문자열일 경우
    if isinstance(s, str):
        try:
            # 문자열은 항상 '%'가 있거나 퍼센트 숫자로 간주하고 100으로 나눔
            return float(s.strip().strip('%')) / 100
        except (ValueError, TypeError):
            # "hello" 같이 변환 불가능한 문자열은 0.0 처리
            return 0.
    # 2. 입력값이 숫자(int, float)일 경우
    elif isinstance(s, (int, float)):
        # 숫자의 절댓값이 1보다 크면 (e.g., 60, -50) 퍼센트로 간주하고 100으로 나눔
        if abs(s) > 1:
            return float(s) / 100
        # 숫자의 절댓값이 1보다 작거나 같으면 (e.g., 0.6, -0.5, 1) 이미 변환된 소수점으로 간주하고 그대로 반환
        else:
            return float(s)

    # 3. 그 외 타입은 0.0 반환
    else:
        return 0.0

def log_validation(message):
    """검증 결과를 세션 상태에 기록합니다."""
    st.session_state.results["validation_log"].append(message)


# =================================================================================================
# --- 외화FS환산용 함수 및 설정 (from fcfs_translate.py) ---
# =================================================================================================
AMOUNT_COL_CANDIDATES = ("외화금액", "FC_Amount", "Amount")
EQUITY_CARRY_COL = "이월금액"
NAME_COL_CANDIDATES = ("계정명", "Account", "Name")
EPS_BS = 1e-6


def _first_numeric_in_row(row):
    s = pd.to_numeric(row, errors="coerce").dropna()
    return None if s.empty else float(s.iloc[0])


def _find_col(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    return None


def read_rates_and_table(xlsx_path):
    all_df = pd.read_excel(xlsx_path, header=0)
    closing_rate = _first_numeric_in_row(all_df.iloc[0])
    average_rate = _first_numeric_in_row(all_df.iloc[1])
    if closing_rate is None or average_rate is None:
        raise ValueError("기말/평균환율을 2~3행(데이터 첫 2행)에서 찾지 못했습니다.")
    df = all_df.drop(index=[0, 1]).reset_index(drop=True)
    if "FS_Element" not in df.columns:
        raise ValueError("파일에 FS_Element 컬럼이 없습니다. (A/L/E/RE/R/X/PI)")
    return closing_rate, average_rate, df


def precheck_foreign_currency(df, eps=EPS_BS):
    df = df.copy()
    amount_col = _find_col(df, AMOUNT_COL_CANDIDATES)
    if amount_col is None:
        raise ValueError(
            f"외화금액 컬럼을 찾지 못했습니다. 후보={AMOUNT_COL_CANDIDATES}"
        )
    df[amount_col] = pd.to_numeric(df[amount_col], errors="coerce").fillna(0.0)
    is_A = df["FS_Element"].eq("A")
    is_L = df["FS_Element"].eq("L")
    is_E = df["FS_Element"].eq("E")
    is_RE = df["FS_Element"].eq("RE")
    is_R = df["FS_Element"].eq("R")
    is_X = df["FS_Element"].eq("X")
    a_fc = df.loc[is_A, amount_col].sum()
    l_fc = df.loc[is_L, amount_col].sum()
    e_fc = df.loc[is_E | is_RE, amount_col].sum()
    ni_fc = df.loc[is_R, amount_col].sum() - df.loc[is_X, amount_col].sum()
    bs_gap_fc = a_fc - l_fc - e_fc
    print(
        f"[PRECHECK] (외화) A-L-(E+RE) = {bs_gap_fc}",
        "->",
        "OK" if abs(bs_gap_fc) < eps else "NG",
    )
    print(f"[PRECHECK] (외화) NI_FC = {ni_fc}")
    return {
        "A_FC": a_fc,
        "L_FC": l_fc,
        "E_plus_RE_FC": e_fc,
        "NI_FC": ni_fc,
        "BS_GAP_FC": bs_gap_fc,
        "BS_OK_FC": abs(bs_gap_fc) < eps,
    }


def translate_fcfs(df, closing_rate, average_rate, eps=EPS_BS):
    df = df.copy()
    amount_col = _find_col(df, AMOUNT_COL_CANDIDATES)
    if amount_col is None:
        raise ValueError(
            f"외화금액 컬럼을 찾지 못했습니다. 후보={AMOUNT_COL_CANDIDATES}"
        )
    name_col = _find_col(df, NAME_COL_CANDIDATES)
    df[amount_col] = pd.to_numeric(df[amount_col], errors="coerce").fillna(0.0)
    if EQUITY_CARRY_COL in df.columns:
        df[EQUITY_CARRY_COL] = pd.to_numeric(
            df[EQUITY_CARRY_COL], errors="coerce"
        ).fillna(0.0)
    out_col = "금액"
    if out_col not in df.columns:
        df[out_col] = 0.0
    is_A = df["FS_Element"].eq("A")
    is_L = df["FS_Element"].eq("L")
    is_E = df["FS_Element"].eq("E")
    is_RE = df["FS_Element"].eq("RE")
    is_R = df["FS_Element"].eq("R")
    is_X = df["FS_Element"].eq("X")
    is_PI = df["FS_Element"].eq("PI")
    df.loc[is_A | is_L, out_col] = df.loc[is_A | is_L, amount_col] * closing_rate
    df.loc[is_E | is_RE, out_col] = (
        df[EQUITY_CARRY_COL] if EQUITY_CARRY_COL in df.columns else 0.0
    )
    df.loc[is_R, out_col] = df.loc[is_R, amount_col] * average_rate
    df.loc[is_X, out_col] = df.loc[is_X, amount_col] * average_rate
    ni_krw = df.loc[is_R, out_col].sum() - df.loc[is_X, out_col].sum()
    if is_RE.any():
        re_idxs = df.index[is_RE]
        df.loc[re_idxs[0], out_col] = df.loc[re_idxs[0], out_col] + ni_krw
    else:
        new_row = {col: None for col in df.columns}
        new_row["FS_Element"] = "RE"
        new_row[out_col] = ni_krw
        new_row[amount_col] = 0.0
        if EQUITY_CARRY_COL in df.columns:
            new_row[EQUITY_CARRY_COL] = 0.0
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
        is_RE = df["FS_Element"].eq("RE")
    assets_sum = df.loc[df["FS_Element"].eq("A"), out_col].sum()
    liabs_sum = df.loc[df["FS_Element"].eq("L"), out_col].sum()
    equity_sum = df.loc[df["FS_Element"].isin(["E", "RE"]), out_col].sum()
    diff = assets_sum - liabs_sum - equity_sum
    if is_PI.any():
        pi_idxs = df.index[is_PI]
        df.loc[pi_idxs, out_col] = 0.0
        df.loc[pi_idxs[0], out_col] = diff
    else:
        new_row = {col: None for col in df.columns}
        new_row["FS_Element"] = "PI"
        new_row[out_col] = diff
        new_row[amount_col] = 0.0
        if EQUITY_CARRY_COL in df.columns:
            new_row[EQUITY_CARRY_COL] = 0.0
        if name_col is not None:
            new_row[name_col] = "해외사업환산손익(PI)"
        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
    pi_krw = diff
    e_total_with_pi = df.loc[df["FS_Element"].isin(["E", "RE", "PI"]), out_col].sum()
    bs_gap_after = assets_sum - liabs_sum - e_total_with_pi
    print(
        f"[POSTCHECK] (환산 후) A-L-(E+RE+PI) = {bs_gap_after:.4f}",
        "  -> ",
        "OK" if abs(bs_gap_after) < eps else "NG",
    )
    print(
        f"[POSTCHECK] A={assets_sum:.2f}, L={liabs_sum:.2f}, (E+RE+PI)={e_total_with_pi:.2f}, NI(from R&X)={ni_krw:.2f}, PI={pi_krw:.2f}"
    )
    totals = {
        "A(KRW)": assets_sum,
        "L(KRW)": liabs_sum,
        "E_plus_RE_plus_PI(KRW)": e_total_with_pi,
        "NI(KRW from R&X)": ni_krw,
        "PI(KRW)": pi_krw,
        "A-L-(E+RE+PI) (after)": bs_gap_after,
    }
    cols_to_check = [amount_col, out_col]
    if EQUITY_CARRY_COL in df.columns:
        cols_to_check.append(EQUITY_CARRY_COL)
    is_zero_row = (df[cols_to_check].fillna(0) == 0).all(axis=1)
    df = df[~is_zero_row].reset_index(drop=True)
    return df, totals


# --- 사이드바 파일 업로드 ---
with st.sidebar:
    st.header("📁 파일 업로드")
    st.info("파일을 업로드하면 세션에 저장됩니다. 세션 종료시 메모리에서 삭제됩니다.")
    st.session_state.files["coa"] = st.file_uploader(
        "1. CoA (계정 체계)", type="xlsx", key="coa_uploader"
    )
    st.session_state.files["parent"] = st.file_uploader(
        "2. 모회사 재무제표 (BSPL, CF 시트 포함)", type="xlsx", key="parent_uploader"
    )
    st.session_state.files["subsidiaries"] = st.file_uploader(
        "3. 자회사 재무제표 (다중 선택 가능)",
        type="xlsx",
        accept_multiple_files=True,
        key="subs_uploader",
    )
    st.session_state.files["adjustment"] = st.file_uploader(
        "4. 연결 조정 분개 (CAJE 업로드)", type="xlsx", key="adj_uploader"
    )

# --- 탭 구성 ---
tab1, tab2, tab3, tab4 = st.tabs(
    ["📈 연결 재무제표", "📝 주석 대사", "🔁 연결조정", "🌍 외화FS환산"]
)

# =================================================================================================
# --- 연결 재무제표 탭 ---
# =================================================================================================
with tab1:
    st.header("1. 연결 재무제표 생성")
    st.write(
        "CoA, 모회사, 자회사 재무제표와 연결 조정 데이터를 통합하여 연결 재무상태표, 손익계산서, 현금흐름표, 자본변동표를 생성합니다."
    )

    def generate_sce_df(coa_df, parent_ce_df, subs_ce_dfs, parent_name, subs_names, adjustment_file, merged_bspl_df):
        """사용자 정의 양식의 CE 시트를 파싱하여 연결 자본변동표(SCE)를 생성합니다."""
        # 1. CoA 기반 동적 컬럼 정의
        e_element_df = coa_df[coa_df['FS_Element'] == 'E'].dropna(axis=1).copy()
        if e_element_df.shape[1] < 4:
            log_validation("⚠️ [자본변동표] CoA의 자본(E) 항목에 레벨 정보가 충분하지 않습니다.")
            return pd.DataFrame()
        level_code_col = e_element_df.columns[-2]
        level_name_col = e_element_df.columns[-1]
        equity_groups = e_element_df[[level_code_col, level_name_col]].dropna().drop_duplicates().sort_values(by=level_code_col)
        l3_codes_map = pd.Series(equity_groups[level_name_col].values, index=equity_groups[level_code_col]).to_dict()

        nci_equity_row = coa_df[coa_df["FS_Element"] == "CE"]
        if not nci_equity_row.empty:
            nci_code = nci_equity_row.iloc[0]["계정코드"]
            nci_name = nci_equity_row.iloc[0]["계정명"]
            l3_codes_map[nci_code] = nci_name
        else:
            log_validation("⚠️ [자본변동표] CoA에서 비지배지분(CE) 계정을 찾을 수 없습니다.")

        sce_cols = list(l3_codes_map.values())
        col_to_l3_map = {v: k for k, v in l3_codes_map.items()}

        # 2. 입력된 CE 시트 파싱 (위치 기반으로 수정)
        all_parsed_dfs = []
        all_input_dfs = [(parent_name, parent_ce_df)] + list(zip(subs_names, subs_ce_dfs))

        for name, df in all_input_dfs:
            if df.empty:
                continue
            try:
                code_row_index = df[df.iloc[:, 2] == '계정코드'].index[0]
                codes = df.iloc[code_row_index, 3:].astype(str).str.strip().tolist()
                data_start_row = code_row_index + 1
                
                data_df = df.iloc[data_start_row:].copy()
                
                num_desc_cols = 3
                num_data_cols = len(codes)
                data_df = data_df.iloc[:, :(num_desc_cols + num_data_cols)]

                # 위치를 기준으로 컬럼 이름을 명시적으로 지정
                desc_names = ['_col_company', '구분', '계정코드'] 
                data_df.columns = desc_names + codes
                data_df['회사명'] = name
                all_parsed_dfs.append(data_df)
            except (IndexError, KeyError) as e:
                log_validation(f"⚠️ [자본변동표] {name}의 CE 시트 양식을 파싱할 수 없습니다: {e}")
                continue

        if not all_parsed_dfs:
            log_validation("⚠️ [자본변동표] 유효한 CE 시트 데이터를 찾을 수 없습니다.")
            return pd.DataFrame()

        combined_ce_df = pd.concat(all_parsed_dfs, ignore_index=True)
        rename_dict = {code: name for code, name in l3_codes_map.items() if code in combined_ce_df.columns}
        combined_ce_df.rename(columns=rename_dict, inplace=True)

        for col in sce_cols:
            if col in combined_ce_df.columns:
                combined_ce_df[col] = pd.to_numeric(combined_ce_df[col], errors='coerce').fillna(0)
            else:
                combined_ce_df[col] = 0

        # 3. 기초자본 계산
        beginning_simple_sum = combined_ce_df[combined_ce_df['계정코드'] == 'Beginning'][sce_cols].sum()

        adj_xls = pd.ExcelFile(adjustment_file)

        if "CAJE_BSPL" in adj_xls.sheet_names:
            full_adj_df = pd.read_excel(adj_xls, "CAJE_BSPL", dtype={'계정코드': str})
        else:
            full_adj_df = pd.DataFrame()

        beginning_adjustments = pd.Series(dtype='float64')
        if not full_adj_df.empty:
            full_adj_df = full_adj_df.dropna(subset=['계정코드'])

            if not full_adj_df.empty:
                full_adj_df['계정코드'] = full_adj_df['계정코드'].astype(str).str.strip().str.split('.').str[0]
                if 'FS_Element' in full_adj_df.columns:
                    full_adj_df = full_adj_df.drop(columns=['FS_Element'])
                full_adj_df = full_adj_df.merge(coa_df[['계정코드', 'FS_Element', 'L3_code']], on='계정코드', how='left')

                # FIX: L3_code가 없는 자본/비지배지분 항목은 계정코드를 L3_code로 사용
                is_equity_like = full_adj_df['FS_Element'].isin(['E', 'CE'])
                is_l3_missing = full_adj_df['L3_code'].isna()
                full_adj_df.loc[is_equity_like & is_l3_missing, 'L3_code'] = full_adj_df.loc[is_equity_like & is_l3_missing, '계정코드']

                full_adj_df['금액'] = pd.to_numeric(full_adj_df['금액'], errors='coerce').fillna(0)

                beg_adj_df = full_adj_df[full_adj_df['당기전기'] != '당기'].copy()
                beg_equity_adjs = beg_adj_df[beg_adj_df['FS_Element'].isin(['E', 'CE'])].copy()

                if not beg_equity_adjs.empty:
                    beg_equity_adjs.loc[:, '금액'] *= -1
                    beginning_adjustments = beg_equity_adjs.groupby('L3_code')['금액'].sum()

        beginning_row = pd.Series(0, index=sce_cols, name='기초')
        beginning_row.update(beginning_simple_sum)
        for code, amount in beginning_adjustments.items():
            if code in l3_codes_map:
                beginning_row[l3_codes_map[code]] += amount

        # 4. 당기 변동분 계산
        current_changes_df = combined_ce_df[~combined_ce_df['계정코드'].isin(['Beginning', 'Ending'])].copy()
        
        r_adj_sum = merged_bspl_df.loc[merged_bspl_df["FS_Element"] == "R", "연결조정"].sum()
        x_adj_sum = merged_bspl_df.loc[merged_bspl_df["FS_Element"] == "X", "연결조정"].sum()
        pl_adj_sum = -r_adj_sum - x_adj_sum
        nci_pl_adj = merged_bspl_df.loc[merged_bspl_df["FS_Element"] == "CR", "연결조정"].sum()
        
        ni_adj_row = pd.DataFrame([{'구분': '당기순손익(연결조정)', '이익잉여금': pl_adj_sum - nci_pl_adj, '비지배지분': nci_pl_adj}]).fillna(0)
        
        curr_adj_df = full_adj_df[full_adj_df['당기전기'] == '당기'].copy()
        curr_equity_adjs = curr_adj_df[curr_adj_df['FS_Element'].isin(['E', 'CE'])].copy()
        curr_equity_adjs.loc[:, '금액'] *= -1
        
        pl_related_codes = merged_bspl_df[merged_bspl_df['FS_Element'].isin(['R', 'X', 'CR'])]['계정코드'].unique()
        direct_equity_adjs = curr_equity_adjs[~curr_equity_adjs['계정코드'].isin(pl_related_codes)]
        direct_adj_by_code = direct_equity_adjs.groupby('L3_code')['금액'].sum()

        direct_adj_row_data = {'구분': '기타자본(연결조정)'}
        for code, amount in direct_adj_by_code.items():
            if code in l3_codes_map:
                direct_adj_row_data[l3_codes_map[code]] = amount
        direct_adj_row = pd.DataFrame([direct_adj_row_data]).fillna(0)

        # 5. 최종 조립
        beg_sce = pd.DataFrame([beginning_row])
        final_sce = pd.concat([beg_sce, current_changes_df.groupby('구분')[sce_cols].sum()], ignore_index=False)
        final_sce = pd.concat([final_sce, ni_adj_row.set_index('구분')], ignore_index=False)
        if not direct_adj_row.empty and direct_adj_row.drop(columns=['구분']).iloc[0].abs().sum() > 1:
             final_sce = pd.concat([final_sce, direct_adj_row.set_index('구분')], ignore_index=False)

        final_sce = final_sce.loc[(final_sce[sce_cols].abs().sum(axis=1)) > 1].fillna(0)
        final_sce.loc['기말', sce_cols] = final_sce[sce_cols].sum()

        # 6. 검증 행 추가
        l3_map = dict(zip(coa_df['계정코드'], coa_df['L3_code']))
        if 'L3_code' not in merged_bspl_df.columns:
             merged_bspl_df['L3_code'] = merged_bspl_df['계정코드'].map(l3_map)

        # For CE elements (NCI), if L3_code is null, use the account code itself.
        is_ce = merged_bspl_df['FS_Element'] == 'CE'
        is_l3_missing = merged_bspl_df['L3_code'].isna()
        merged_bspl_df.loc[is_ce & is_l3_missing, 'L3_code'] = merged_bspl_df.loc[is_ce & is_l3_missing, '계정코드']

        l3_totals = merged_bspl_df.groupby('L3_code')['연결금액'].sum()
        
        verification_row = pd.Series(index=sce_cols, name="검증(연결BS)")
        for col, code in col_to_l3_map.items():
            verification_row[col] = l3_totals.get(code, 0)
        final_sce.loc['검증(연결BS)'] = verification_row

        final_sce = final_sce.reset_index().rename(columns={'index': '구분'})
        
        # '구분'에 중복이 있을 수 있으므로, 첫 번째 '계정코드'를 사용하도록 중복을 제거하여 map을 생성
        temp_map_df = combined_ce_df[['구분', '계정코드']].dropna(subset=['구분']).drop_duplicates(subset=['구분'])
        row_to_code_map = pd.Series(temp_map_df.계정코드.values, index=temp_map_df.구분).to_dict()

        row_to_code_map.update({'기초': 'Beginning', '기말': 'Ending', '검증(연결BS)': 'Verification', '당기순손익(연결조정)': 'CE11_NI', '기타자본(연결조정)': 'CE12_CAJE'})
        final_sce.insert(1, '조정코드', final_sce['구분'].map(row_to_code_map).fillna('CE9999'))
        
        return final_sce

    if st.button(
        "🚀 연결 재무제표 생성 실행",
        key="run_consolidation",
        disabled=not (
            st.session_state.files["coa"] and st.session_state.files["parent"]
        ),
    ):
        with st.spinner("데이터를 처리하고 있습니다... 잠시만 기다려주세요."):
            # Reset previous results
            st.session_state.results["validation_log"] = []
            st.session_state.results["consolidation_wp_bs"] = None
            st.session_state.results["consolidation_wp_pl"] = None
            st.session_state.results["consolidation_wp_cf"] = None
            st.session_state.results["consolidation_wp_sce"] = None


            # 파일명에서 회사 이름 추출
            parent_name = st.session_state.files["parent"].name.split("_")[0]
            subs_names = [
                f.name.split("_")[0] for f in st.session_state.files["subsidiaries"]
            ]

            try:
                # ----------------------------------------------------------------
                # 1. 데이터 준비 (파일 읽기 및 전처리)
                # ----------------------------------------------------------------
                @st.cache_data
                def load_and_clean_data(
                    coa_file, parent_file, parent_name, subs_files, subs_names, adj_file
                ):
                    def clean_df(df, key_col="계정코드"):
                        if key_col in df.columns:
                            df[key_col] = (
                                df[key_col]
                                .astype(str)
                                .str.strip()
                                .str.split(".")
                                .str[0]
                            )
                            df = df.dropna(subset=[key_col])
                        return df

                    def read_fs_sheets(file, file_name=""):
                        try:
                            file.seek(0)
                            xls = pd.ExcelFile(file)
                            bspl_df = (
                                pd.read_excel(
                                    xls, sheet_name="BSPL", dtype={"계정코드": str}
                                )
                                if "BSPL" in xls.sheet_names
                                else pd.DataFrame()
                            )
                            cf_df = (
                                pd.read_excel(
                                    xls,
                                    sheet_name="CF",
                                    dtype={"계정코드": str, "CF_code": str},
                                )
                                if "CF" in xls.sheet_names
                                else pd.DataFrame()
                            )

                            bspl_df = clean_df(bspl_df, "계정코드")
                            if "CF_code" in cf_df.columns:
                                cf_df = clean_df(cf_df, "CF_code")
                            elif "계정코드" in cf_df.columns:
                                cf_df = clean_df(cf_df, "계정코드").rename(
                                    columns={"계정코드": "CF_code"}
                                )

                            for df in [bspl_df, cf_df]:
                                if "금액" in df.columns:
                                    df["금액"] = pd.to_numeric(
                                        df["금액"], errors="coerce"
                                    ).fillna(0)

                            return bspl_df, cf_df
                        except Exception as e:
                            st.error(f"'{file_name}' 파일 처리 중 오류: {e}")
                            return pd.DataFrame(), pd.DataFrame()
                    
                    coa_file.seek(0)
                    coa_df = clean_df(
                        pd.read_excel(coa_file, sheet_name="CoA", dtype=str), "계정코드"
                    )
                    xls_coa = pd.ExcelFile(coa_file)
                    cf_coa_df = pd.DataFrame()
                    if "CF" in xls_coa.sheet_names:
                        cf_coa_df = pd.read_excel(xls_coa, sheet_name="CF", dtype=str)
                        if "CF_code" in cf_coa_df.columns:
                            cf_coa_df = clean_df(cf_coa_df, "CF_code")
                    else:
                        log_validation(
                            "경고: CoA 파일에 'CF' 시트가 없습니다. 현금흐름표 집계가 제한될 수 있습니다."
                        )

                    aje_code = pd.read_excel(coa_file, sheet_name="AJE", dtype=str)

                    parent_bspl_df, parent_cf_df = read_fs_sheets(
                        parent_file, parent_name
                    )
                    parent_bspl_df = parent_bspl_df.rename(
                        columns={"금액": parent_name}
                    )
                    parent_cf_df = parent_cf_df.rename(columns={"금액": parent_name})

                    subs_bspl_dfs, subs_cf_dfs = [], []
                    for f, sub_name in zip(subs_files, subs_names):
                        bspl, cf = read_fs_sheets(f, sub_name)
                        subs_bspl_dfs.append(bspl.rename(columns={"금액": sub_name}))
                        subs_cf_dfs.append(cf.rename(columns={"금액": sub_name}))

                    caje_bspl_df, caje_cf_df, re_code = (
                        pd.DataFrame(),
                        pd.DataFrame(),
                        None,
                    )
                    if adj_file:
                        try:
                            adj_file.seek(0)
                            xls_adj = pd.ExcelFile(adj_file)
                            if "CAJE_BSPL" in xls_adj.sheet_names:
                                caje_bspl_df = pd.read_excel(
                                    xls_adj,
                                    sheet_name="CAJE_BSPL",
                                    dtype={"계정코드": str},
                                )
                                caje_bspl_df = clean_df(caje_bspl_df, "계정코드")
                                if "금액" in caje_bspl_df.columns:
                                    caje_bspl_df["금액"] = pd.to_numeric(
                                        caje_bspl_df["금액"], errors="coerce"
                                    ).fillna(0)

                            if "CAJE_CF" in xls_adj.sheet_names:
                                caje_cf_df = pd.read_excel(
                                    xls_adj,
                                    sheet_name="CAJE_CF",
                                    dtype={"계정코드": str},
                                )
                                caje_cf_df = clean_df(caje_cf_df, "계정코드")
                                if "조정금액" in caje_cf_df.columns:
                                    caje_cf_df["조정금액"] = pd.to_numeric(
                                        caje_cf_df["조정금액"], errors="coerce"
                                    ).fillna(0)




                        except Exception as e:
                            log_validation(
                                f"🚨 오류: 조정분개 파일({adj_file.name}) 처리 중 오류 발생: {e}"
                            )

                    return (
                        coa_df,
                        cf_coa_df,
                        parent_bspl_df,
                        parent_cf_df,
                        subs_bspl_dfs,
                        subs_cf_dfs,
                        caje_bspl_df,
                        caje_cf_df,
                        re_code,
                        aje_code,
                    )

                (
                    coa_df,
                    cf_coa_df,
                    parent_bspl_df,
                    parent_cf_df,
                    subs_bspl_dfs,
                    subs_cf_dfs,
                    caje_bspl_df,
                    caje_cf_df_from_file,
                    re_code,
                    aje_code
                ) = load_and_clean_data(
                    st.session_state.files["coa"],
                    st.session_state.files["parent"],
                    parent_name,
                    st.session_state.files["subsidiaries"],
                    tuple(subs_names),  # 리스트는 해시 불가능하므로 튜플로 변환
                    st.session_state.files["adjustment"],
                )

                # ----------------------------------------------------------------
                # 2. 데이터 검증
                # ----------------------------------------------------------------
                def check_duplicates(df, name):
                    if "계정코드" in df.columns:
                        dups = df["계정코드"].value_counts().loc[lambda x: x > 1]
                        if not dups.empty:
                            log_validation(
                                f"⚠️ **[{name}]** 중복 계정코드 발견: {', '.join(dups.index)}"
                            )

                def check_missing_in_coa(df, coa_codes, name):
                    if "계정코드" in df.columns:
                        missing = set(df["계정코드"]) - coa_codes
                        if missing:
                            log_validation(
                                f"🚨 **[{name}]** CoA에 없는 계정코드 발견: {', '.join(sorted(list(missing)))}"
                            )

                def check_balance_sheet_equation(df, coa_df, column_name):
                    """재무상태표 차대 검증 (자산 = 부채 + 자본)"""
                    if "계정코드" in df.columns and column_name in df.columns:
                        if "FS_Element" in df.columns:
                            merged = df
                        elif "FS_Element" in coa_df.columns:
                            merged = df.merge(
                                coa_df[["계정코드", "FS_Element"]],
                                on="계정코드",
                                how="left",
                            )
                        else:
                            return  # Cannot perform check

                        total_assets = pd.to_numeric(
                            merged[merged["FS_Element"].isin(["A", "CA"])][column_name],
                            errors="coerce",
                        ).sum()
                        total_liabilities = pd.to_numeric(
                            merged[merged["FS_Element"] == "L"][column_name],
                            errors="coerce",
                        ).sum()
                        total_equity = pd.to_numeric(
                            merged[merged["FS_Element"].isin(["E", "CE"])][column_name],
                            errors="coerce",
                        ).sum()
                        difference = total_assets - (total_liabilities + total_equity)

                        if abs(difference) > 1:  # 사소한 반올림 오류는 무시
                            log_validation(
                                f"❌ **[{column_name}]** 재무상태표 차대 불일치: {difference:,.0f}"
                            )
                        else:
                            log_validation(
                                f"✅ **[{column_name}]** 재무상태표 차대 일치"
                            )

                check_duplicates(parent_bspl_df, parent_name)
                for name, df in zip(subs_names, subs_bspl_dfs):
                    check_duplicates(df, name)

                coa_codes = set(coa_df["계정코드"])
                check_missing_in_coa(parent_bspl_df, coa_codes, parent_name)
                for name, df in zip(subs_names, subs_bspl_dfs):
                    check_missing_in_coa(df, coa_codes, name)

                # ----------------------------------------------------------------
                # 3. BS/PL 데이터 통합 및 계산
                # ----------------------------------------------------------------
                merged_bspl_df = coa_df.merge(parent_bspl_df[["계정코드", parent_name]], on="계정코드", how="left", sort=False)
                for name, df in zip(subs_names, subs_bspl_dfs):
                    merged_bspl_df = merged_bspl_df.merge(df[["계정코드", name]], on="계정코드", how="left", sort=False)

                bspl_val_cols = [parent_name] + subs_names
                merged_bspl_df[bspl_val_cols] = merged_bspl_df[bspl_val_cols].fillna(0)
                merged_bspl_df["단순합계"] = merged_bspl_df[bspl_val_cols].sum(axis=1)

                check_balance_sheet_equation(merged_bspl_df, coa_df, parent_name)
                for name in subs_names:
                    check_balance_sheet_equation(merged_bspl_df, coa_df, name)
                check_balance_sheet_equation(merged_bspl_df, coa_df, "단순합계")

                if not caje_bspl_df.empty and "계정코드" in caje_bspl_df.columns:
                    adj_bspl_grouped = caje_bspl_df.groupby("계정코드")["금액"].sum().reset_index()
                    adj_with_fs = adj_bspl_grouped.merge(coa_df[["계정코드", "FS_Element"]], on="계정코드", how="left")
                    is_ler = adj_with_fs["FS_Element"].isin(["L", "E", "R", "CE", "CR"])
                    adj_with_fs.loc[is_ler, "금액"] *= -1
                    merged_bspl_df = merged_bspl_df.merge(adj_with_fs[["계정코드", "금액"]].rename(columns={"금액": "연결조정"}), on="계정코드", how="left", sort=False)
                    merged_bspl_df["연결조정"] = merged_bspl_df["연결조정"].fillna(0)
                else:
                    merged_bspl_df["연결조정"] = 0

                
                nci_equity_row = coa_df[coa_df["FS_Element"] == "CE"]
                if not nci_equity_row.empty:
                    nci_code = nci_equity_row.iloc[0]["계정코드"]
                    nci_name = nci_equity_row.iloc[0]["계정명"]
                
                nci_pl_row = coa_df[coa_df["FS_Element"] == "CR"]
                if not nci_pl_row.empty:
                    NCI_PL_CODE = nci_pl_row.iloc[0]["계정코드"]
                    NCI_PL_NAME = nci_pl_row.iloc[0]["계정명"]

                nci_pl_sum = merged_bspl_df.loc[merged_bspl_df["FS_Element"] == "CR", "연결조정"].sum()

                if not nci_pl_row.empty and not nci_equity_row.empty:
                    nci_equity_code = nci_equity_row.iloc[0]['계정코드']
                    if (merged_bspl_df['계정코드'] == nci_equity_code).any():
                        merged_bspl_df.loc[merged_bspl_df['계정코드'] == nci_equity_code, '연결조정'] += nci_pl_sum
                
                r_adj_sum = merged_bspl_df.loc[merged_bspl_df["FS_Element"] == "R", "연결조정"].sum()
                x_adj_sum = merged_bspl_df.loc[merged_bspl_df["FS_Element"] == "X", "연결조정"].sum()
                pl_adj_sum = r_adj_sum - x_adj_sum

                aje_code = pd.read_excel(
                    st.session_state.files["coa"], sheet_name="AJE", dtype=str
                )
                re_code = aje_code.loc[aje_code["FS_Element"] == "E", "계정코드"].iloc[0]
                merged_bspl_df.loc[merged_bspl_df["계정코드"] == re_code, "연결조정"] += pl_adj_sum
                        
                merged_bspl_df["연결금액"] = merged_bspl_df["단순합계"] + merged_bspl_df["연결조정"]
                log_validation("--- 연결금액 기준 차대 검증 ---")
                check_balance_sheet_equation(merged_bspl_df, coa_df, "연결금액")

                # ----------------------------------------------------------------
                # 4. CF 데이터 통합 및 계산
                # ----------------------------------------------------------------
                CF_KEY = "CF_code"
                merged_cf_df = pd.DataFrame()
                if not cf_coa_df.empty and CF_KEY in cf_coa_df.columns:
                    merged_cf_df = cf_coa_df.merge(parent_cf_df[[CF_KEY, parent_name]], on=CF_KEY, how="left", sort=False)
                    for name, df in zip(subs_names, subs_cf_dfs):
                        if CF_KEY in df.columns:
                            merged_cf_df = merged_cf_df.merge(df[[CF_KEY, name]], on=CF_KEY, how="left", sort=False)
                    
                    cf_val_cols = [parent_name] + subs_names
                    merged_cf_df[cf_val_cols] = merged_cf_df[cf_val_cols].fillna(0)
                    merged_cf_df["단순합계"] = merged_cf_df[cf_val_cols].sum(axis=1)

                    if not caje_cf_df_from_file.empty and "계정코드" in caje_cf_df_from_file.columns and "조정금액" in caje_cf_df_from_file.columns:
                        adj_cf_grouped = caje_cf_df_from_file.groupby("계정코드")["조정금액"].sum().reset_index()
                        merged_cf_df = merged_cf_df.merge(adj_cf_grouped.rename(columns={"조정금액": "연결조정"}), on="계정코드", how="left")
                    else:
                        merged_cf_df["연결조정"] = 0
                    
                    merged_cf_df["연결조정"] = merged_cf_df["연결조정"].fillna(0)
                    merged_cf_df["연결금액"] = merged_cf_df["단순합계"] + merged_cf_df["연결조정"]

                # ----------------------------------------------------------------
                # 5. 소계 및 최종 FS 생성
                # ----------------------------------------------------------------
                def generate_fs_with_subtotals(
                    df,
                    name_cols,
                    amount_cols,
                    name_code_map,
                    desc_col="계정명",
                    code_col="계정코드",
                ):
                    df = df.copy()

                    # Sign 로직 적용
                    apply_sign_logic = "sign" in df.columns
                    if apply_sign_logic:
                        df["sign"] = pd.to_numeric(df["sign"], errors="coerce").fillna(
                            1
                        )
                        for col in amount_cols:
                            if col in df.columns:
                                df[col] = df[col] * df["sign"]

                    # 소계 계산을 위한 재귀 함수
                    def recursive_subtotal(data, current_name_cols, level=0):
                        if not current_name_cols or data.empty:
                            return data

                        current_col, remaining_cols = (
                            current_name_cols[0],
                            current_name_cols[1:],
                        )
                        all_sub_dfs = []

                        # 레벨 정보가 있는 그룹 먼저 처리
                        for key, group in data.dropna(subset=[current_col]).groupby(
                            current_col, sort=False
                        ):
                            sub_df = recursive_subtotal(
                                group, remaining_cols, level + 1
                            )

                            # 합계 행 생성
                            sum_data = {col: "" for col in data.columns}
                            sum_data.update(group[amount_cols].sum())
                            sum_data[desc_col] = f"{'' * level}{key}"  # 들여쓰기
                            sum_data[code_col] = name_code_map.get(key, "")

                            # FS_Element, sign 등 메타데이터 복사
                            if not group.empty:
                                for col in ["FS_Element", "sign"]:
                                    if col in group.columns:
                                        sum_data[col] = group.iloc[0][col]

                            sum_row = pd.DataFrame([sum_data])
                            all_sub_dfs.append(
                                pd.concat([sub_df, sum_row], ignore_index=True)
                            )

                        # 레벨 정보가 없는(NaN) 그룹을 나중에 처리하여 아래로 보냄
                        nan_group = data[data[current_col].isna()]
                        if not nan_group.empty:
                            all_sub_dfs.append(
                                recursive_subtotal(nan_group, remaining_cols, level + 1)
                            )

                        # all_sub_dfs가 비어있는 경우 에러 방지
                        if not all_sub_dfs:
                            return pd.DataFrame(columns=data.columns)

                        return pd.concat(all_sub_dfs, ignore_index=True)

                    final_df = recursive_subtotal(df, name_cols)

                    # Sign 원복
                    if apply_sign_logic and not final_df.empty:
                        final_df["sign"] = (
                            pd.to_numeric(final_df["sign"], errors="coerce")
                            .fillna(1)
                            .replace(0, 1)
                        )
                        final_df[amount_cols] = final_df[amount_cols].divide(
                            final_df["sign"], axis=0
                        )

                    return final_df

                # BS, PL, CF 데이터 분리 및 소계 생성
                df_bs = merged_bspl_df[
                    merged_bspl_df["FS_Element"].isin(["A", "L", "E", "CA", "CE"])
                ].copy()
                df_bs["sign"] = (
                    df_bs["FS_Element"]
                    .map({"A": -1, "CA": -1, "L": 1, "E": 1, "CE": 1})
                    .fillna(1)
                )

                df_pl = merged_bspl_df[
                    merged_bspl_df["FS_Element"].isin(["R", "X", "CR"])
                ].copy()
                df_pl["sign"] = df_pl["FS_Element"].map({"R": 1, "X": -1, "CR": 1}).fillna(1)

                df_cf = merged_cf_df.copy()
                if "FS_Element" in df_cf.columns:  # CF의 FS_Element는 부호로 사용
                    df_cf["sign"] = pd.to_numeric(
                        df_cf["FS_Element"], errors="coerce"
                    ).fillna(1)

                # 소계 생성을 위한 설정
                con_amtcols = (
                    [parent_name] + subs_names + ["단순합계", "연결조정", "연결금액"]
                )
                bspl_name_cols = [
                    c
                    for c in coa_df.columns
                    if c.startswith("L") and not c.endswith("code")
                ]
                cf_name_cols = [
                    c
                    for c in cf_coa_df.columns
                    if c.startswith("L") and not c.endswith("code")
                ]

                # 이름-코드 매핑 생성
                bspl_name_code_map = {
                    row[name]: row[code]
                    for code, name in zip(
                        [
                            c
                            for c in coa_df.columns
                            if c.startswith("L") and c.endswith("code")
                        ],
                        bspl_name_cols,
                    )
                    for _, row in coa_df.iterrows()
                    if pd.notna(row.get(name))
                }
                cf_name_code_map = {
                    row[name]: row[code]
                    for code, name in zip(
                        [
                            c
                            for c in cf_coa_df.columns
                            if c.startswith("L") and c.endswith("code")
                        ],
                        cf_name_cols,
                    )
                    for _, row in cf_coa_df.iterrows()
                    if pd.notna(row.get(name))
                }

                # 최종 FS 생성
                bs_final = generate_fs_with_subtotals(
                    df_bs, bspl_name_cols, con_amtcols, bspl_name_code_map
                )
                pl_final = generate_fs_with_subtotals(
                    df_pl, bspl_name_cols, con_amtcols, bspl_name_code_map
                )
                cf_final = generate_fs_with_subtotals(
                    df_cf,
                    cf_name_cols,
                    con_amtcols,
                    cf_name_code_map,
                    desc_col="현금흐름표",
                    code_col="CF_code",
                )

                # 불필요한 레벨 컬럼 제거 및 최종 정리
                level_cols = [c for c in coa_df.columns if c.startswith("L")] + [
                    c for c in cf_coa_df.columns if c.startswith("L")
                ]
                l_cols_to_drop = list(set(level_cols + ["sign"]))
                bs_final.drop(
                    columns=[c for c in l_cols_to_drop if c in bs_final.columns],
                    inplace=True,
                )
                pl_final.drop(
                    columns=[c for c in l_cols_to_drop if c in pl_final.columns],
                    inplace=True,
                )
                cf_final.drop(
                    columns=[c for c in l_cols_to_drop if c in cf_final.columns],
                    inplace=True,
                )

                # 소계 행 식별 및 'is_subtotal' 컬럼 추가
                bspl_name_cols = [
                    c
                    for c in coa_df.columns
                    if c.startswith("L") and not c.endswith("code")
                ]
                if bspl_name_cols:
                    bspl_subtotal_names = set(coa_df[bspl_name_cols].stack().unique())
                    if not bs_final.empty:
                        bs_final["is_subtotal"] = bs_final["계정명"].isin(
                            bspl_subtotal_names
                        )
                    if not pl_final.empty:
                        pl_final["is_subtotal"] = pl_final["계정명"].isin(
                            bspl_subtotal_names
                        )

                cf_name_cols = [
                    c
                    for c in cf_coa_df.columns
                    if c.startswith("L") and not c.endswith("code")
                ]
                if not cf_coa_df.empty and cf_name_cols:
                    cf_subtotal_names = set(cf_coa_df[cf_name_cols].stack().unique())
                    if not cf_final.empty:
                        cf_final["is_subtotal"] = cf_final["현금흐름표"].isin(
                            cf_subtotal_names
                        )

                # 0에 가까운 값 정리 및 정수 변환
                processed_dfs = []
                for df in [bs_final, pl_final, cf_final]:
                    if not df.empty:
                        df = df.copy()
                        amt_cols_in_df = [c for c in con_amtcols if c in df.columns]

                        if amt_cols_in_df:
                            df.loc[
                                (df[amt_cols_in_df].abs().sum(axis=1)) < 0.01,
                                amt_cols_in_df,
                            ] = 0
                            df[amt_cols_in_df] = (
                                df[amt_cols_in_df].fillna(0).round().astype("int64")
                            )

                            # 금액이 모두 0이면서 소계가 아닌 행을 제거
                            if "is_subtotal" in df.columns:
                                all_zeros = (df[amt_cols_in_df] == 0).all(axis=1)
                                is_not_subtotal = df["is_subtotal"] == False
                                rows_to_remove = all_zeros & is_not_subtotal
                                df = df[~rows_to_remove]
                            else:
                                log_validation(
                                    "⚠️ 경고: 'is_subtotal' 컬럼이 없어 일부 0원 행이 제거되지 않을 수 있습니다."
                                )
                    processed_dfs.append(df)
                bs_final, pl_final, cf_final = processed_dfs

                # --- CF 검증 로직 추가 (연결조정 합계) ---
                if (
                    not df_cf.empty
                    and "연결조정" in df_cf.columns
                    and "sign" in df_cf.columns
                ):
                    # Sign을 반영한 연결조정 금액의 합계가 0인지 검증
                    total_cf_adjustment = (df_cf["연결조정"] * df_cf["sign"]).sum()
                    if abs(total_cf_adjustment) > 1:  # 사소한 반올림 오류는 무시
                        log_validation(
                            f"❌ **[현금흐름표 검증]** 연결조정의 합계(부호 반영)가 0이 아닙니다: {total_cf_adjustment:,.0f}"
                        )
                    else:
                        log_validation(
                            f"✅ **[현금흐름표 검증]** 연결조정의 합계(부호 반영)가 0으로 일치합니다."
                        )

                # CF 두번째 행(당기순이익 부분합) 제거
                cf_final = cf_final.drop(cf_final.index[1])
                # 세션 상태에 결과 저장
                st.session_state.results["consolidation_wp_bs"] = bs_final
                st.session_state.results["consolidation_wp_pl"] = pl_final
                st.session_state.results["consolidation_wp_cf"] = cf_final

                # --- 6. 자본변동표 생성 ---
                sce_final = pd.DataFrame()
                try:
                    parent_file = st.session_state.files["parent"]
                    parent_file.seek(0)
                    xls_parent = pd.ExcelFile(parent_file)
                    parent_ce_df = pd.read_excel(xls_parent, sheet_name="CE", header=None) if "CE" in xls_parent.sheet_names else pd.DataFrame()

                    subs_ce_dfs = []
                    for f in st.session_state.files["subsidiaries"]:
                        f.seek(0)
                        xls_sub = pd.ExcelFile(f)
                        subs_ce_dfs.append(pd.read_excel(xls_sub, sheet_name="CE", header=None) if "CE" in xls_sub.sheet_names else pd.DataFrame())
                    
                    adj_file = st.session_state.files["adjustment"]
                    if adj_file:
                        adj_file.seek(0)
                        sce_final = generate_sce_df(coa_df, parent_ce_df, subs_ce_dfs, parent_name, subs_names, adj_file, merged_bspl_df)
                    else:
                        log_validation("⚠️ [자본변동표] 조정분개 파일이 없어 자본변동표를 생성할 수 없습니다.")

                except Exception as e:
                    log_validation(f"⚠️ **[자본변동표 생성 오류]** 자본변동표 생성 중 오류가 발생했습니다: {e}")
                
                st.session_state.results["consolidation_wp_sce"] = sce_final
                # --------------------------

                st.success("🎉 연결 재무제표 생성이 완료되었습니다!")

            except Exception as e:
                st.error(f"연결 재무제표 생성 중 오류가 발생했습니다: {e}")
                st.exception(e)

    # --- 결과 표시 ---
    if st.session_state.results["validation_log"]:
        with st.expander("🔍 데이터 검증 로그 보기", expanded=True):
            for log in st.session_state.results["validation_log"]:
                st.markdown(log, unsafe_allow_html=True)

    if (
        st.session_state.results.get("consolidation_wp_bs") is not None
        and not st.session_state.results["consolidation_wp_bs"].empty
    ):
        st.subheader("📄 연결 재무상태표")
        st.dataframe(st.session_state.results["consolidation_wp_bs"].style.format(precision=0, thousands=","))
        st.subheader("📄 연결 손익계산서")
        st.dataframe(st.session_state.results["consolidation_wp_pl"].style.format(precision=0, thousands=","))
        st.subheader("📄 연결 현금흐름표")
        st.dataframe(st.session_state.results["consolidation_wp_cf"].style.format(precision=0, thousands=","))
        
        if st.session_state.results.get("consolidation_wp_sce") is not None and not st.session_state.results["consolidation_wp_sce"].empty:
            st.subheader("📄 연결 자본변동표")
            st.dataframe(st.session_state.results["consolidation_wp_sce"].style.format(precision=0, thousands=","))

        # --- 다운로드 버튼 ---
        excel_data = to_excel({
                "Consol_BS": st.session_state.results["consolidation_wp_bs"],
                "Consol_PL": st.session_state.results["consolidation_wp_pl"],
                "Consol_CF": st.session_state.results["consolidation_wp_cf"],
                "Consol_SCE": st.session_state.results.get("consolidation_wp_sce", pd.DataFrame()),
            })
        st.download_button(
            label="📥 전체 결과 다운로드 (Excel)",
            data=excel_data,
            file_name="consolidated_fs_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    elif not (st.session_state.files["coa"] and st.session_state.files["parent"]):
        st.info("사이드바에서 CoA와 모회사 자회사 연결조정분개 파일을 업로드한 후 '생성 실행' 버튼을 눌러주세요.")

with tab2:
    st.header("2. 주석 대사 (Reconciliation)")
    st.write(
        "모회사 주석을 기준으로 자회사 주석들의 숫자 데이터를 위치 기반으로 합산하고, 연결정산표와 대사합니다."
    )
    footnote_parent_file = st.file_uploader("1. 모회사 주석 파일", type="xlsx")
    footnote_subs_files = st.file_uploader(
        "2. 자회사 주석 파일 (다중 선택 가능)", type="xlsx", accept_multiple_files=True
    )
    if st.button("🔄 주석 대사 실행", disabled=not footnote_parent_file):
        if (
            st.session_state.results.get("consolidation_wp_bs") is None
            and footnote_subs_files
        ):
            st.warning(
                "대사를 위해서는 먼저 '연결 재무제표' 탭에서 '생성 실행'을 완료해야 합니다."
            )
            st.stop()
        with st.spinner("주석 파일을 취합하고 대사하고 있습니다..."):
            try:
                st.session_state.results["combined_footnotes"] = {}
                parent_sheets = pd.read_excel(
                    footnote_parent_file, sheet_name=None, dtype=str
                )
                subs_files_data = [
                    (Path(f.name).stem, pd.read_excel(f, sheet_name=None, dtype=str))
                    for f in footnote_subs_files
                ]

                conso_wp_df = pd.concat(
                    [
                        st.session_state.results.get(
                            "consolidation_wp_bs", pd.DataFrame()
                        ),
                        st.session_state.results.get(
                            "consolidation_wp_pl", pd.DataFrame()
                        ),
                    ]
                )
                conso_map = (
                    conso_wp_df.set_index("계정코드")["연결금액"].to_dict()
                    if not conso_wp_df.empty
                    else {}
                )

                for sheet_name, parent_df in parent_sheets.items():
                    if "주석" not in sheet_name:
                        continue
                    all_dfs_for_sheet = []
                    parent_df_copy = parent_df.copy()
                    parent_df_copy["소스파일"] = Path(footnote_parent_file.name).stem
                    all_dfs_for_sheet.append(parent_df_copy)
                    for name, sheets in subs_files_data:
                        if sheet_name in sheets:
                            sub_df_copy = sheets[sheet_name].copy()
                            sub_df_copy["소스파일"] = name
                            all_dfs_for_sheet.append(sub_df_copy)
                    should_concat = any(
                        pd.to_numeric(df[col], errors="coerce").isna().any()
                        for df in all_dfs_for_sheet
                        for col in df.columns[2:-1]
                        if len(df.columns) > 3
                    )
                    if should_concat:
                        final_df = pd.concat(all_dfs_for_sheet, ignore_index=True)
                    else:
                        final_df = all_dfs_for_sheet[0].copy()
                        value_cols = final_df.columns[2:-1]
                        final_df[value_cols] = (
                            final_df[value_cols]
                            .apply(pd.to_numeric, errors="coerce")
                            .fillna(0)
                        )
                        for next_df in all_dfs_for_sheet[1:]:
                            next_value_cols = next_df.columns[2:-1]
                            next_df[next_value_cols] = (
                                next_df[next_value_cols]
                                .apply(pd.to_numeric, errors="coerce")
                                .fillna(0)
                            )
                            final_df[value_cols] = final_df[value_cols].add(
                                next_df[next_value_cols].values, fill_value=0
                            )
                        if footnote_subs_files:
                            final_df["소스파일"] = "combined"
                        if "계정코드" in final_df.columns and not conso_wp_df.empty:
                            numeric_cols = final_df.select_dtypes(
                                include="number"
                            ).columns
                            if not numeric_cols.empty:
                                last_numeric_col = numeric_cols[-1]
                                if "연결조정" in conso_wp_df.columns:
                                    adj_map = conso_wp_df.set_index("계정코드")[
                                        "연결조정"
                                    ].to_dict()
                                    final_df["계정코드_str"] = (
                                        final_df["계정코드"].astype(str).str.strip()
                                    )
                                    adj_values = (
                                        final_df["계정코드_str"].map(adj_map).fillna(0)
                                    )
                                    final_df[last_numeric_col] += adj_values
                                    final_df = final_df.drop(columns=["계정코드_str"])
                        if "계정코드" in final_df.columns and conso_map:
                            last_numeric_col = final_df.select_dtypes(
                                include="number"
                            ).columns[-1]

                            def check_value_match(row):
                                code = str(row["계정코드"]).strip()
                                if not code:
                                    return ""
                                footnote_value = row[last_numeric_col]
                                conso_value = conso_map.get(code)
                                if conso_value is None:
                                    return "불일치 (정산표에 코드 없음)"
                                if abs(footnote_value - conso_value) < 1:
                                    return "일치"
                                else:
                                    return f"불일치 (차이: {footnote_value - conso_value:,.0f})"

                            final_df["대사결과"] = final_df.apply(
                                check_value_match, axis=1
                            )
                    st.session_state.results["combined_footnotes"][
                        sheet_name
                    ] = final_df
                st.success("🎉 주석 취합 및 대사가 완료되었습니다!")
            except Exception as e:
                st.error(f"주석 처리 중 오류가 발생했습니다: {e}")

    if st.session_state.results.get("combined_footnotes"):
        st.subheader("📒 취합된 주석 데이터")
        for sheet_name, df in st.session_state.results["combined_footnotes"].items():
            with st.expander(f"시트: {sheet_name}", expanded=False):
                st.dataframe(df)
        footnote_excel_data = to_excel(st.session_state.results["combined_footnotes"])
        st.download_button(
            label="📥 취합된 주석 다운로드 (Excel)",
            data=footnote_excel_data,
            file_name="combined_footnotes.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

with tab3:
    st.header("3. 연결 조정 분개 생성")
    st.write(
        "기본 조정명세 입력 후, 법인세/비지배지분(NCI) 조정명세를 자동 생성하고, 최종 검토 후 전체 조정분개를 생성합니다."
    )

    # --- Session State for Tab 3 ---
    if "adj_workflow" not in st.session_state:
        st.session_state.adj_workflow = {
            "initial_file": None,
            "intermediate_data": None,
            "final_file": None,
        }

    @st.cache_data
    def create_adjustment_template():
        adjustment_types = [
            "CAJE00_투자자본상계",
            "CAJE01_채권채무제거",
            "CAJE02_제품미실현이익제거",
            "CAJE03_상각자산미실현이익제거",
            "CAJE04_배당조정",
            "CAJE05_기타손익조정",
            "CAJE96_취득일차이조정",
            "CAJE97_법인세조정",
            "CAJE98_비지배지분조정",
            "CAJE99_기타조정",
        ]
        columns = ["회사명", "계정코드", "계정명", "당기전기", "금액", "설명"]
        output = io.BytesIO()

        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            # 'Info' 시트 생성
            info_data = {
                "당기세율": [0.2, 0.18, 0.16],
                "당기지분율": [1, 0.60, 0.80],
            }
            info_index_labels = ["모회사", "자회사A", "자회사B"]
            info_df = pd.DataFrame(info_data, index=info_index_labels)
            info_df.index.name = "회사명"
            info_df.to_excel(writer, sheet_name="Info")

            # 'Info' 시트 스타일 적용
            ws_info = writer.sheets["Info"]
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(
                start_color="4F81BD", end_color="4F81BD", fill_type="solid"
            )
            header_alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            for cell in ws_info[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            ws_info.column_dimensions[get_column_letter(1)].width = 20
            for i, col_name in enumerate(info_df.columns, 2):
                ws_info.column_dimensions[get_column_letter(i)].width = 20

            for sheet_name in adjustment_types:
                if sheet_name == "CAJE00_투자자본상계":
                    example_data = [
                        {
                            "회사명": "모회사",
                            "계정코드": "19200",
                            "계정명": "종속기업투자",
                            "당기전기": "취득일",
                            "금액": 3400000,
                            "설명": "자회사A 투자금액 제거",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "33100",
                            "계정명": "자본금",
                            "당기전기": "취득일",
                            "금액": 3000000,
                            "설명": "자회사A 자본금 제거",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "37500",
                            "계정명": "이익잉여금",
                            "당기전기": "취득일",
                            "금액": 1000000,
                            "설명": "자회사A 이익잉여금(취득시점) 제거",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "20600",
                            "계정명": "기계장치",
                            "당기전기": "취득일",
                            "금액": -800000,
                            "설명": "자회사A 공정가치차이 계상",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "101000",
                            "계정명": "영업권",
                            "당기전기": "취득일",
                            "금액": -200000,
                            "설명": "자회사A 영업권 계상",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "201100",
                            "계정명": "비지배지분",
                            "당기전기": "취득일",
                            "금액": -1600000,
                            "설명": "자회사A 비지배지분 계상",
                        },
                    ]
                    df = pd.DataFrame(example_data)
                elif sheet_name == "CAJE01_채권채무제거":
                    example_data = [
                        {
                            "회사명": "모회사",
                            "계정코드": "10800",
                            "계정명": "매출채권",
                            "당기전기": "전기",
                            "금액": 20000000,
                            "설명": "자회사A에 대한 매출채권 제거",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "25100",
                            "계정명": "매입채무",
                            "당기전기": "전기",
                            "금액": 20000000,
                            "설명": "모회사에 대한 매입채무 제거",
                        },
                        {
                            "회사명": "모회사",
                            "계정코드": "10800",
                            "계정명": "매출채권",
                            "당기전기": "당기",
                            "금액": 10000000,
                            "설명": "자회사A에 대한 매출채권 제거",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "25100",
                            "계정명": "매입채무",
                            "당기전기": "당기",
                            "금액": 10000000,
                            "설명": "모회사에 대한 매입채무 제거",
                        },
                    ]
                    df = pd.DataFrame(example_data)
                elif sheet_name == "CAJE02_제품미실현이익제거":
                    example_data = [
                        {
                            "회사명": "모회사",
                            "계정코드": "45500",
                            "계정명": "매출원가",
                            "당기전기": "전기",
                            "금액": 3000000,
                            "설명": "전기 미실현이익",
                        },
                        {
                            "회사명": "모회사",
                            "계정코드": "37500",
                            "계정명": "이익잉여금",
                            "당기전기": "전기",
                            "금액": 3000000,
                            "설명": "전기 미실현이익(이익잉여금)",
                        },
                        {
                            "회사명": "모회사",
                            "계정코드": "40200",
                            "계정명": "매출",
                            "당기전기": "당기",
                            "금액": 10000000,
                            "설명": "당기 판매분 매출",
                        },
                        {
                            "회사명": "모회사",
                            "계정코드": "45500",
                            "계정명": "매출원가",
                            "당기전기": "당기",
                            "금액": 6000000,
                            "설명": "당기 판매분 매출원가",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "15200",
                            "계정명": "제품(재고자산)",
                            "당기전기": "당기",
                            "금액": 4000000,
                            "설명": "모회사가 판매한 재고 미실현이익 제거(재고감소)",
                        },
                    ]
                    df = pd.DataFrame(example_data)
                elif sheet_name == "CAJE03_상각자산미실현이익제거":
                    example_data = [
                        {
                            "회사명": "모회사",
                            "계정코드": "20600",
                            "계정명": "기계장치",
                            "당기전기": "전기",
                            "금액": 5000000,
                            "설명": "자회사A에서 모회사에 처분 모회사 보유",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "37500",
                            "계정명": "이익잉여금",
                            "당기전기": "전기",
                            "금액": 5000000,
                            "설명": "자회사A 계상 유형자산처분이익",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "37500",
                            "계정명": "이익잉여금",
                            "당기전기": "전기",
                            "금액": -1000000,
                            "설명": "감가상각비 증분 제거 - 유형자산처분이익 효과 감소",
                        },
                        {
                            "회사명": "모회사",
                            "계정코드": "20700",
                            "계정명": "기계장치감가상각누계액",
                            "당기전기": "전기",
                            "금액": -1000000,
                            "설명": "감가상각누계액 증가",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "81800",
                            "계정명": "감가상각비",
                            "당기전기": "당기",
                            "금액": 1000000,
                            "설명": "감가상각비 증분 제거 - 유형자산처분이익 효과 감소",
                        },
                        {
                            "회사명": "모회사",
                            "계정코드": "20700",
                            "계정명": "기계장치감가상각누계액",
                            "당기전기": "당기",
                            "금액": -1000000,
                            "설명": "감가상각누계액 증가",
                        },
                    ]
                    df = pd.DataFrame(example_data)
                elif sheet_name == "CAJE04_배당조정":
                    example_data = [
                        {"회사명": "모회사", "계정코드": "90300", "계정명": "배당금수익", "당기전기": "당기", "금액": 1200000, "설명": "자회사A로부터 받은 배당금수익 제거"},
                        {"회사명": "자회사A", "계정코드": "201100", "계정명": "비지배지분", "당기전기": "당기", "금액": 800000, "설명": "배당금 비지배지분 조정"},
                        {"회사명": "자회사A", "계정코드": "37500", "계정명": "이익잉여금", "당기전기": "당기", "금액": -2000000, "설명": "모회사에 지급한 배당금 효과 제거"},
                    ]
                    df = pd.DataFrame(example_data)
                elif sheet_name == "CAJE96_취득일차이조정":
                    example_data = [
                        {
                            "회사명": "자회사A",
                            "계정코드": "37500",
                            "계정명": "이익잉여금",
                            "당기전기": "전기",
                            "금액": 160000,
                            "설명": "자회사A 공정가치차이 상각",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "20700",
                            "계정명": "기계장치감가상각누계액",
                            "당기전기": "전기",
                            "금액": 160000,
                            "설명": "자회사A 공정가치차이 상각누계액",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "81800",
                            "계정명": "감가상각비",
                            "당기전기": "당기",
                            "금액": -160000,
                            "설명": "자회사A 공정가치차이 상각",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "20700",
                            "계정명": "기계장치감가상각누계액",
                            "당기전기": "당기",
                            "금액": 160000,
                            "설명": "자회사A 공정가치차이 상각누계액",
                        },
                    ]
                    df = pd.DataFrame(example_data)
                elif sheet_name == "CAJE96_취득일차이조정":
                    example_data = [
                        {
                            "회사명": "자회사A",
                            "계정코드": "37500",
                            "계정명": "이익잉여금",
                            "당기전기": "전기",
                            "금액": 160000,
                            "설명": "자회사A 공정가치차이 상각",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "20700",
                            "계정명": "기계장치감가상각누계액",
                            "당기전기": "전기",
                            "금액": 160000,
                            "설명": "자회사A 공정가치차이 상각누계액",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "81800",
                            "계정명": "감가상각비",
                            "당기전기": "당기",
                            "금액": -160000,
                            "설명": "자회사A 공정가치차이 상각",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "20700",
                            "계정명": "기계장치감가상각누계액",
                            "당기전기": "당기",
                            "금액": 160000,
                            "설명": "자회사A 공정가치차이 상각누계액",
                        },
                    ]
                    df = pd.DataFrame(example_data)
                elif sheet_name == "CAJE97_법인세조정":
                    example_data = [
                        {
                            "회사명": "자회사A",
                            "계정코드": "37500",
                            "계정명": "이월이익잉여금",
                            "당기전기": "취득일",
                            "금액": 144000,
                            "설명": "취득일 공정가치차이 법인세 효과",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "31000",
                            "계정명": "이연법인세부채",
                            "당기전기": "취득일",
                            "금액": -144000,
                            "설명": "취득일 공정가치차이 법인세 효과",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "37500",
                            "계정명": "이월이익잉여금",
                            "당기전기": "전기",
                            "금액": -28800,
                            "설명": "취득일 공정가치차이 상각 법인세 효과",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "31000",
                            "계정명": "이연법인세부채",
                            "당기전기": "전기",
                            "금액": 28800,
                            "설명": "취득일 공정가치차이 상각 법인세 효과",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "99800",
                            "계정명": "법인세비용",
                            "당기전기": "전기",
                            "금액": -600000,
                            "설명": "재고 미실현이익 실현 법인세 효과",
                        },
                        {
                            "회사명": "모회사",
                            "계정코드": "37500",
                            "계정명": "이월이익잉여금",
                            "당기전기": "전기",
                            "금액": -600000,
                            "설명": "재고 미실현이익 실현 법인세 효과",
                        },
                        {
                            "회사명": "모회사",
                            "계정코드": "37500",
                            "계정명": "이익잉여금",
                            "당기전기": "전기",
                            "금액": -800000,
                            "설명": "유형자산처분이익 효과 감소 관련 법인세효과",
                        },
                        {
                            "회사명": "모회사",
                            "계정코드": "31000",
                            "계정명": "이연법인세부채",
                            "당기전기": "전기",
                            "금액": 800000,
                            "설명": "유형자산처분이익 효과 감소 관련 법인세효과",
                        },
                    ]
                    df = pd.DataFrame(example_data)
                elif sheet_name == "CAJE98_비지배지분조정":
                    example_data = [
                        {
                            "회사명": "자회사A",
                            "계정코드": "37500",
                            "계정명": "이월이익잉여금",
                            "당기전기": "전기",
                            "금액": 500000,
                            "설명": "전기 누적 자본변동 비지배지분",
                        },
                        {
                            "회사명": "자회사A",
                            "계정코드": "201100",
                            "계정명": "비지배지분",
                            "당기전기": "전기",
                            "금액": -500000,
                            "설명": "전기 누적 자본변동 비지배지분",
                        },
                    ]
                    df = pd.DataFrame(example_data)
                else:
                    df = pd.DataFrame(columns=columns)
                df = df.reindex(columns=columns)
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                ws = writer.sheets[sheet_name]
                header_font = Font(bold=True, color="FFFFFF")
                header_fill = PatternFill(
                    start_color="4F81BD", end_color="4F81BD", fill_type="solid"
                )
                header_alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                for i, column_name in enumerate(df.columns, 1):
                    ws.column_dimensions[get_column_letter(i)].width = 20

            # --- 데이터 유효성 검사 추가 ---
            num_companies = len(info_index_labels)
            validation_formula = f"='Info'!$A$2:$A${num_companies + 1}"
            dv = DataValidation(
                type="list", formula1=validation_formula, allow_blank=True
            )
            dv.error = "목록에 있는 값만 입력할 수 있습니다."
            dv.errorTitle = "잘못된 입력"
            dv.prompt = "목록에서 회사명을 선택하세요."
            dv.promptTitle = "회사명 선택"

            target_range = "A2:A100001"
            for sheet_name in adjustment_types:
                if sheet_name != "Info":
                    ws = writer.sheets[sheet_name]
                    ws.add_data_validation(dv)
                    dv.add(target_range)
        return output.getvalue()

    def generate_intermediate_adjustments(adj_file, coa_df, subs_files, subs_names, aje_code):
        adj_file.seek(0)
        xls = pd.ExcelFile(adj_file)
        original_sheets = {
            sheet_name: pd.read_excel(xls, sheet_name, dtype={"계정코드": str}) for sheet_name in xls.sheet_names
        }

        if "Info" not in original_sheets:
            st.error("'Info' 시트가 조정분개 파일에 없습니다.")
            return None

        info_df = original_sheets["Info"].set_index(original_sheets["Info"].columns[0])

        info_df["당기세율_num"] = info_df["당기세율"].apply(parse_percent)
        info_df["당기지분율_num"] = info_df["당기지분율"].apply(parse_percent)
        tax_rates = info_df["당기세율_num"].to_dict()
        nci_rates = (1 - info_df["당기지분율_num"]).to_dict()

        # Create maps for faster lookups
        fs_map = dict(zip(coa_df["계정코드"].astype(str), coa_df["FS_Element"]))
        name_map = dict(zip(coa_df["계정코드"].astype(str), coa_df["계정명"]))
        tax_adj_entries, nci_adj_entries = [], []

        # Get special account codes from aje_code DataFrame and CoA
        nci_pl_row = coa_df[coa_df["FS_Element"] == "CR"]
        if not nci_pl_row.empty:
            NCI_PL_CODE = nci_pl_row.iloc[0]["계정코드"]
            NCI_PL_NAME = nci_pl_row.iloc[0]["계정명"]
        else:
            NCI_PL_CODE = "302000"  # Fallback
            NCI_PL_NAME = "비지배지분순손익"
            st.warning("CoA 파일에서 'CR' FS_Element를 가진 비지배지분순손익 계정을 찾을 수 없습니다. 기본값('302000')을 사용합니다.")

        nci_equity_row = coa_df[coa_df["FS_Element"] == "CE"]
        if not nci_equity_row.empty:
            NCI_EQUITY_CODE = nci_equity_row.iloc[0]["계정코드"]
            NCI_EQUITY_NAME = nci_equity_row.iloc[0]["계정명"]
        else:
            NCI_EQUITY_CODE = "201100"  # Fallback
            NCI_EQUITY_NAME = "비지배지분"
            st.warning("CoA 파일에서 'CE' FS_Element를 가진 비지배지분 계정을 찾을 수 없습니다. 기본값('201100')을 사용합니다.")

        IT_EXPENSE_CODE = aje_code.loc[aje_code["FS_Element"] == "X", "계정코드"].iloc[0]
        IT_EXPENSE_NAME = aje_code.loc[aje_code["FS_Element"] == "X", "계정명"].iloc[0]
        DTA_CODE = aje_code.loc[aje_code["FS_Element"] == "L", "계정코드"].iloc[0]
        DTA_NAME = aje_code.loc[aje_code["FS_Element"] == "L", "계정명"].iloc[0]
        RE_CODE = aje_code.loc[aje_code["FS_Element"] == "E", "계정코드"].iloc[0]
        RE_NAME = aje_code.loc[aje_code["FS_Element"] == "E", "계정명"].iloc[0]

        # --- 1. Tax and NCI on P/L adjustments from CAJE sheets ---
        for sheet_name, df_orig in original_sheets.items():
            sheet_name_upper = sheet_name.upper()
            if not sheet_name_upper.startswith("CAJE"):
                continue

            caje_type = sheet_name_upper.split("_")[0]
            if caje_type in ["CAJE97", "CAJE98"]:
                continue

            df = df_orig[df_orig["당기전기"] == "당기"].copy()
            if df.empty:
                continue

            df["금액"] = pd.to_numeric(df.get("금액"), errors="coerce")
            df = df.dropna(subset=["금액", "계정코드", "회사명"])
            if df.empty:
                continue
            df["계정코드"] = df["계정코드"].astype(str).str.strip()
            df["FS_Element"] = df["계정코드"].map(fs_map)

            if caje_type == "CAJE02":
                pl_rows = df[df["FS_Element"].isin(["R", "X"])]

                df_orig_with_fs = df_orig.copy()
                df_orig_with_fs["계정코드"] = (
                    df_orig_with_fs["계정코드"].astype(str).str.strip()
                )
                df_orig_with_fs["FS_Element"] = df_orig_with_fs["계정코드"].map(fs_map)
                asset_rows = df_orig_with_fs[df_orig_with_fs["FS_Element"] == "A"]

                if asset_rows.empty or pl_rows.empty:
                    continue

                asset_corp = asset_rows.iloc[0]["회사명"]
                tax_rate = tax_rates.get(asset_corp, 0.0)

                # Tax effect
                unrealized_profit_amount = asset_rows["금액"].sum()
                tax_effect = unrealized_profit_amount * tax_rate
                if abs(tax_effect) > 1:
                    desc = f"[{sheet_name}] 미실현이익 법인세효과"
                    tax_adj_entries.append(
                        {
                            "회사명": asset_corp,
                            "계정코드": IT_EXPENSE_CODE,
                            "계정명": IT_EXPENSE_NAME,
                            "당기전기": "당기",
                            "금액": tax_effect,
                            "설명": desc,
                        }
                    )
                    tax_adj_entries.append(
                        {
                            "회사명": asset_corp,
                            "계정코드": DTA_CODE,
                            "계정명": DTA_NAME,
                            "당기전기": "당기",
                            "금액": tax_effect,
                            "설명": desc,
                        }
                    )

                # NCI effect
                for _, pl_row in pl_rows.iterrows():
                    pl_corp = pl_row["회사명"]
                    nci_rate = nci_rates.get(pl_corp, 0.0)
                    if nci_rate > 0:
                        amount = pl_row["금액"]
                        nci_effect = -amount * (1 - tax_rate) * nci_rate
                        if abs(nci_effect) > 1:
                            desc = f"[{sheet_name}] {pl_row.get('설명', '')} 관련"
                            nci_adj_entries.append(
                                {
                                    "회사명": asset_corp,
                                    "계정코드": RE_CODE,
                                    "계정명": RE_NAME,
                                    "당기전기": "당기",
                                    "금액": -nci_effect,
                                    "설명": f"{desc} 비지배지분효과",
                                }
                            )
                            nci_adj_entries.append(
                                {
                                    "회사명": asset_corp,
                                    "계정코드": NCI_PL_CODE,
                                    "계정명": NCI_PL_NAME,
                                    "당기전기": "당기",
                                    "금액": nci_effect,
                                    "설명": f"{desc} 비지배지분효과",
                                }
                            )

            elif caje_type == "CAJE03":
                pl_rows = df[df["FS_Element"].isin(["R", "X"])]

                df_orig_with_fs = df_orig.copy()
                df_orig_with_fs["계정코드"] = (
                    df_orig_with_fs["계정코드"].astype(str).str.strip()
                )
                df_orig_with_fs["FS_Element"] = df_orig_with_fs["계정코드"].map(fs_map)
                asset_rows = df_orig_with_fs[df_orig_with_fs["FS_Element"] == "A"]

                if asset_rows.empty or pl_rows.empty:
                    continue

                asset_corp = asset_rows.iloc[0]["회사명"]

                for _, pl_row in pl_rows.iterrows():
                    pl_corp = pl_row["회사명"]
                    amount = pl_row["금액"]
                    tax_rate = tax_rates.get(asset_corp, 0.0)
                    desc = f"[{sheet_name}] {pl_row.get('설명', '')} 관련"

                    # Tax effect
                    tax_effect = -amount * tax_rate
                    if abs(tax_effect) > 1:
                        tax_adj_entries.append(
                            {
                                "회사명": asset_corp,
                                "계정코드": IT_EXPENSE_CODE,
                                "계정명": IT_EXPENSE_NAME,
                                "당기전기": "당기",
                                "금액": tax_effect,
                                "설명": f"{desc} 법인세효과",
                            }
                        )
                        tax_adj_entries.append(
                            {
                                "회사명": asset_corp,
                                "계정코드": DTA_CODE,
                                "계정명": DTA_NAME,
                                "당기전기": "당기",
                                "금액": tax_effect,
                                "설명": f"{desc} 법인세효과",
                            }
                        )
                        
                    nci_rate = nci_rates.get(pl_corp, 0.0)

                    # NCI effect
                    if nci_rate > 0:
                        nci_effect = -amount * (1 - tax_rate) * nci_rate
                        if abs(nci_effect) > 1:
                            nci_adj_entries.append(
                                {
                                    "회사명": asset_corp,
                                    "계정코드": RE_CODE,
                                    "계정명": RE_NAME,
                                    "당기전기": "당기",
                                    "금액": -nci_effect,
                                    "설명": f"{desc} 비지배지분효과",
                                }
                            )
                            nci_adj_entries.append(
                                {
                                    "회사명": asset_corp,
                                    "계정코드": NCI_PL_CODE,
                                    "계정명": NCI_PL_NAME,
                                    "당기전기": "당기",
                                    "금액": nci_effect,
                                    "설명": f"{desc} 비지배지분효과",
                                }
                            )

            elif caje_type == "CAJE96":
                pl_rows = df[df["FS_Element"].isin(["R", "X"])]
                for _, pl_row in pl_rows.iterrows():
                    corp = pl_row["회사명"]
                    amount = pl_row["금액"]
                    tax_rate = tax_rates.get(corp, 0.0)
                    nci_rate = nci_rates.get(corp, 0.0)
                    desc = f"[{sheet_name}] {pl_row.get('설명', '')} 관련"

                    # Tax effect
                    tax_effect = -amount * tax_rate
                    if abs(tax_effect) > 1:
                        tax_adj_entries.append(
                            {
                                "회사명": corp,
                                "계정코드": IT_EXPENSE_CODE,
                                "계정명": IT_EXPENSE_NAME,
                                "당기전기": "당기",
                                "금액": tax_effect,
                                "설명": f"{desc} 법인세효과",
                            }
                        )
                        tax_adj_entries.append(
                            {
                                "회사명": corp,
                                "계정코드": DTA_CODE,
                                "계정명": DTA_NAME,
                                "당기전기": "당기",
                                "금액": tax_effect,
                                "설명": f"{desc} 법인세효과",
                            }
                        )

                    # NCI effect
                    if nci_rate > 0:
                        nci_effect = -amount * (1 - tax_rate) * nci_rate
                        if abs(nci_effect) > 1:
                            nci_adj_entries.append(
                                {
                                    "회사명": corp,
                                    "계정코드": RE_CODE,
                                    "계정명": RE_NAME,
                                    "당기전기": "당기",
                                    "금액": -nci_effect,
                                    "설명": f"{desc} 비지배지분효과",
                                }
                            )
                            nci_adj_entries.append(
                                {
                                    "회사명": corp,
                                    "계정코드": NCI_PL_CODE,
                                    "계정명": NCI_PL_NAME,
                                    "당기전기": "당기",
                                    "금액": nci_effect,
                                    "설명": f"{desc} 비지배지분효과",
                                }
                            )

            elif caje_type == "CAJE05":
                pl_rows = df[df["FS_Element"].isin(["R", "X"])]
                for _, pl_row in pl_rows.iterrows():
                    corp = pl_row["회사명"]
                    amount = pl_row["금액"]
                    tax_rate = tax_rates.get(corp, 0.0)
                    desc = f"[{sheet_name}] {pl_row.get('설명', '')} 관련"
                    tax_effect = -amount * tax_rate
                    if abs(tax_effect) > 1:
                        tax_adj_entries.append(
                            {
                                "회사명": corp,
                                "계정코드": IT_EXPENSE_CODE,
                                "계정명": IT_EXPENSE_NAME,
                                "당기전기": "당기",
                                "금액": tax_effect,
                                "설명": f"{desc} 법인세효과",
                            }
                        )
                        tax_adj_entries.append(
                            {
                                "회사명": corp,
                                "계정코드": DTA_CODE,
                                "계정명": DTA_NAME,
                                "당기전기": "당기",
                                "금액": tax_effect,
                                "설명": f"{desc} 법인세효과",
                            }
                        )

        # --- 2. NCI on subsidiary's total equity change from 'CE' sheet ---
        for sub_file, sub_name in zip(subs_files, subs_names):
            try:
                sub_file.seek(0)
                sub_xls = pd.ExcelFile(sub_file)
                if "CE" in sub_xls.sheet_names:
                    sce_df = pd.read_excel(sub_xls, sheet_name="CE", header=None)

                    if sce_df.shape[0] < 5 or sce_df.shape[1] < 4:
                        continue

                    nci_rate = nci_rates.get(sub_name, 0.0)
                    if nci_rate <= 0:
                        continue

                    for r in range(3, len(sce_df)):
                        data_row = sce_df.iloc[r]
                        row_desc = str(data_row.iloc[1])

                        if "기말" in row_desc or "Ending" in row_desc:
                            break
                        if "기초" in row_desc or "Beginning" in row_desc:
                            continue

                        ce_adj_code = str(data_row.iloc[2])
                        is_ni_item = "_NI" in ce_adj_code
                        nci_contra_code = NCI_PL_CODE if is_ni_item else NCI_EQUITY_CODE
                        nci_contra_name = NCI_PL_NAME if is_ni_item else NCI_EQUITY_NAME

                        # Get the numeric values for the current row
                        change_values = pd.to_numeric(
                            data_row.iloc[3:-1], errors="coerce"
                        ).fillna(0)

                        # Calculate the sum for the row (안분 대상 합계)
                        row_sum = change_values.sum()

                        # If there's no significant change in the row, skip to the next row
                        if abs(row_sum) <= 1:
                            continue

                        # Calculate the total NCI amount for the row
                        total_nci_per_row = row_sum * nci_rate

                        # Create a safe divisor to avoid ZeroDivisionError
                        safe_row_sum = row_sum if row_sum != 0 else 1

                        # Calculate the weight of each column's value relative to the row sum
                        weights = change_values / safe_row_sum

                        # Distribute the total NCI amount across the columns based on weights
                        nci_distribution = weights * total_nci_per_row

                        # Create adjustment entries for each distributed NCI amount
                        header_row = sce_df.iloc[1]
                        equity_acct_codes = header_row.iloc[3:-1]
                        for i, nci_effect in enumerate(nci_distribution):
                            if abs(nci_effect) > 1:
                                equity_acct_code = str(equity_acct_codes.iloc[i])

                                # Debit entry to the specific equity account
                                nci_adj_entries.append(
                                    {
                                        "회사명": sub_name,
                                        "계정코드": equity_acct_code,
                                        "계정명": name_map.get(equity_acct_code, ""),
                                        "당기전기": "당기",
                                        "금액": nci_effect,
                                        "설명": f"{sub_name} 자본변동 ({row_desc})",
                                    }
                                )
                                # Credit entry to the NCI contra account
                                nci_adj_entries.append(
                                    {
                                        "회사명": sub_name,
                                        "계정코드": nci_contra_code,
                                        "계정명": nci_contra_name,
                                        "당기전기": "당기",
                                        "금액": -nci_effect,
                                        "설명": f"{sub_name} 자본변동 ({row_desc})",
                                    }
                                )

                else:
                    log_validation(f"⚠️ **[{sub_name}]** 자본변동표(CE) 시트가 없어 자본변동에 따른 비지배지분 조정을 계산할 수 없습니다.")
            except Exception as e:
                st.warning(f"{sub_name}의 'CE'시트 처리 중 오류: {e}")

        # --- Final assembly of adjustment sheets ---
        final_sheets = original_sheets.copy()
        sheet_names = list(final_sheets.keys())

        def find_sheet_name_by_prefix(prefix):
            for name in sheet_names:
                if name.upper().startswith(prefix):
                    return name
            return f"{prefix}_자동생성" # Fallback to a new sheet name

        # Handle Tax Adjustments (CAJE97)
        caje97_sheet_name = find_sheet_name_by_prefix("CAJE97")
        new_tax_df = pd.DataFrame(tax_adj_entries)
        if caje97_sheet_name in final_sheets and not final_sheets[caje97_sheet_name].empty:
            original_tax_df = final_sheets[caje97_sheet_name].dropna(how="all")
            final_sheets[caje97_sheet_name] = pd.concat(
                [original_tax_df, new_tax_df], ignore_index=True
            )
        elif not new_tax_df.empty:
            final_sheets[caje97_sheet_name] = new_tax_df

        # Handle NCI Adjustments (CAJE98)
        caje98_sheet_name = find_sheet_name_by_prefix("CAJE98")
        new_nci_df = pd.DataFrame(nci_adj_entries)
        if caje98_sheet_name in final_sheets and not final_sheets[caje98_sheet_name].empty:
            original_nci_df = final_sheets[caje98_sheet_name].dropna(how="all")
            final_sheets[caje98_sheet_name] = pd.concat(
                [original_nci_df, new_nci_df], ignore_index=True
            )
        elif not new_nci_df.empty:
            final_sheets[caje98_sheet_name] = new_nci_df

        return to_excel(final_sheets)


    # --- Step 1: Download Template ---
    st.subheader("Step 1: 템플릿 다운로드")
    st.write(
        "템플릿을 다운로드하여 기본 조정 명세서 시트를 작성합니다."
    )
    template_data = create_adjustment_template()
    st.download_button(
        label="📥 조정명세 입력 템플릿 다운로드 (.xlsx)",
        data=template_data,
        file_name="조정명세_입력템플릿_BeforeTaxNci.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # --- Step 2: Upload Initial Adjustments ---
    st.subheader("Step 2: 기본 조정 파일 업로드")
    st.session_state.adj_workflow["initial_file"] = st.file_uploader(
        "작성한 기본 조정 파일을 업로드하세요. ('Info' 시트 포함)",
        type="xlsx",
        key="initial_adj_uploader",
    )

    # --- Step 3: Generate & Download Intermediate File ---
    st.subheader("Step 3: 법인세/NCI 자동계산 및 검토")
    if st.button(
        "⚙️ 법인세/NCI 조정 자동계산 실행",
        disabled=not (
            st.session_state.adj_workflow.get("initial_file")
            and st.session_state.files.get("coa")
        ),
    ):
        st.session_state.results["validation_log"] = []  # 로그 초기화

        subs_files = st.session_state.files.get("subsidiaries", [])
        if not subs_files:
            log_validation(
                "⚠️ **[자회사 파일 없음]** 자회사 재무제표 파일이 업로드되지 않으면 자본변동에 따른 비지배지분 조정을 계산할 수 없습니다."
            )

        with st.spinner("자동 조정 분개를 생성 중입니다..."):
            try:
                coa_df = pd.read_excel(
                    st.session_state.files["coa"], sheet_name="CoA", dtype=str
                )
                subs_names = [f.name.split("_")[0] for f in subs_files]
                aje_code = pd.read_excel(
                    st.session_state.files["coa"], sheet_name="AJE", dtype=str
                )
                intermediate_excel_data = generate_intermediate_adjustments(
                    st.session_state.adj_workflow["initial_file"],
                    coa_df,
                    subs_files,
                    subs_names,
                    aje_code,
                )
                if intermediate_excel_data:
                    st.session_state.adj_workflow["intermediate_data"] = (
                        intermediate_excel_data
                    )
                    st.success(
                        "자동계산이 완료되었습니다. 아래 버튼으로 파일을 다운로드하여 검토하세요."
                    )
            except Exception as e:
                st.error(f"자동계산 중 오류 발생: {e}")
                st.exception(e)

    if st.session_state.results["validation_log"]:
        with st.expander("🔍 조정 자동계산 검증 로그", expanded=True):
            for log in st.session_state.results["validation_log"]:
                st.markdown(log, unsafe_allow_html=True)

    if st.session_state.adj_workflow.get("intermediate_data"):
        st.download_button(
            label="📥 검토용 파일 다운로드 (자동계산 포함)",
            data=st.session_state.adj_workflow["intermediate_data"],
            file_name="조정명세_입력템플릿_TaxNci.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    # --- Step 4: Upload Final File ---
    st.subheader("Step 4: 최종 조정 파일 업로드")
    st.write("검토 및 수정을 완료한 최종 조정 파일을 업로드합니다.")
    st.session_state.adj_workflow["final_file"] = st.file_uploader(
        "최종 조정명세 파일을 업로드하세요.", type="xlsx", key="final_adj_uploader"
    )

    # --- Step 5 & 6: Generate Final CAJE and Display ---
    st.subheader("Step 5: 최종 분개 생성 및 결과 확인")

    def build_caje_from_template(adjustment_file, coa_df_internal):
        fs_map = dict(zip(coa_df_internal["계정코드"], coa_df_internal["FS_Element"]))

        def get_bspl_sign(fs_element):
            return -1 if fs_element in ["A", "X", "CA"] else 1

        ni_code = None
        try:
            r_rows = coa_df_internal[coa_df_internal["FS_Element"] == "R"]
            if not r_rows.empty:
                ni_code = r_rows.iloc[0].get("L1_code")
        except (IndexError, KeyError):
            ni_code = None

        xls = pd.ExcelFile(adjustment_file)
        all_bspl_entries, all_cf_entries = [], []

        for sheet_name in xls.sheet_names:
            if not sheet_name.upper().startswith("CAJE"):
                continue
            caje_type = sheet_name.split("_")[0].upper()
            df = pd.read_excel(xls, sheet_name, dtype={"계정코드": str}).fillna("")

            # --- A. BS/PL Adjustment Logic ---
            df_for_bspl = df.copy()
            if caje_type in ["CAJE01", "CAJE04"]:
                df_for_bspl = df[df["당기전기"] == "당기"]

            for _, row in df_for_bspl.iterrows():
                acc_code = str(row.get("계정코드", "")).strip()
                if not acc_code:
                    continue
                fs_element = fs_map.get(acc_code, "")
                amount = pd.to_numeric(row.get("금액"), errors="coerce")
                if pd.isna(amount) or amount == 0:
                    continue

                final_amount = amount * get_bspl_sign(fs_element)
                all_bspl_entries.append(
                    {
                        "조정유형": caje_type,
                        "회사명": row.get("회사명"),
                        "계정코드": acc_code,
                        "금액": final_amount,
                        "설명": row.get("설명"),
                        "당기전기": row.get("당기전기"),
                        "FS_Element": fs_element,
                    }
                )

            # --- B. CF Adjustment Logic ---
            if caje_type == "CAJE02":
                if ni_code is None:
                    st.warning(
                        f"[{sheet_name}] CF조정 건너뜀: 당기순이익 계정 코드를 CoA에서 찾을 수 없습니다."
                    )
                    continue
                df_with_fs = df.merge(
                    coa_df_internal[["계정코드", "FS_Element"]],
                    on="계정코드",
                    how="left",
                )
                pl_rows = df_with_fs[df_with_fs["FS_Element"].isin(["R", "X"])].copy()
                bs_rows = df_with_fs[
                    df_with_fs["FS_Element"].isin(["A", "L", "E"])
                ].copy()
                if pl_rows.empty or bs_rows.empty:
                    st.warning(
                        f"[{sheet_name}] CF조정 건너뜀: 시트에서 손익(R/X) 또는 재무상태(A/L/E) 계정을 찾을 수 없습니다."
                    )
                    continue

                pl_pivot = pl_rows.pivot_table(
                    index=["계정코드", "FS_Element"],
                    columns="당기전기",
                    values="금액",
                    aggfunc="sum",
                ).fillna(0)
                if "당기" not in pl_pivot.columns:
                    pl_pivot["당기"] = 0
                if "전기" not in pl_pivot.columns:
                    pl_pivot["전기"] = 0
                pl_pivot["change"] = pl_pivot["당기"] + pl_pivot["전기"]
                pl_pivot["impact"] = pl_pivot.apply(
                    lambda r: r["change"] if r.name[1] == "X" else -r["change"], axis=1
                )
                total_pl_impact = pl_pivot["impact"].sum()

                inventory_acc_code = bs_rows.iloc[0]["계정코드"]
                corp_name = df.iloc[0]["회사명"]

                # Line 1: NI Entry (+)
                all_cf_entries.append(
                    {
                        "조정유형": caje_type,
                        "회사명": corp_name,
                        "계정코드": ni_code,
                        "조정금액": total_pl_impact,
                        "설명": "[비현금손익] 미실현이익(NI)",
                    }
                )
                # Line 2: Inventory Entry (-)
                all_cf_entries.append(
                    {
                        "조정유형": caje_type,
                        "회사명": corp_name,
                        "계정코드": inventory_acc_code,
                        "조정금액": -total_pl_impact,
                        "설명": "[비현금손익] 미실현이익(재고)",
                    }
                )
            elif caje_type == "CAJE03":
                if ni_code is None:
                    st.warning(
                        f"[{sheet_name}] CF조정 건너뜀: 당기순이익 계정 코드를 CoA에서 찾을 수 없습니다."
                    )
                    continue
                df_with_fs = df.merge(
                    coa_df_internal[["계정코드", "FS_Element"]],
                    on="계정코드",
                    how="left",
                )
                pl_rows = df_with_fs[df_with_fs["FS_Element"].isin(["X", "R"])].copy()
                bs_rows = df_with_fs[
                    df_with_fs["FS_Element"].isin(["A", "L", "E"])
                ].copy()
                if pl_rows.empty or bs_rows.empty:
                    st.warning(
                        f"[{sheet_name}] CF조정 건너뜀: 시트에서 손익(R/X) 또는 재무상태(A/L/E) 계정을 찾을 수 없습니다."
                    )
                    continue

                pl_pivot = pl_rows.pivot_table(
                    index=["계정코드", "FS_Element"],
                    columns="당기전기",
                    values="금액",
                    aggfunc="sum",
                ).fillna(0)
                if "당기" not in pl_pivot.columns:
                    pl_pivot["당기"] = 0
                if "전기" not in pl_pivot.columns:
                    pl_pivot["전기"] = 0
                pl_pivot["change"] = pl_pivot["당기"] + pl_pivot["전기"]
                pl_pivot["impact"] = pl_pivot.apply(
                    lambda r: r["change"] if r.name[1] == "X" else -r["change"], axis=1
                )
                total_pl_impact = pl_pivot["impact"].sum()

                pl_acc_code = pl_rows.iloc[0]["계정코드"]
                corp_name = df.iloc[0]["회사명"]

                # Line 1: NI Entry (+)
                all_cf_entries.append(
                    {
                        "조정유형": caje_type,
                        "회사명": corp_name,
                        "계정코드": ni_code,
                        "조정금액": total_pl_impact,
                        "설명": "[비현금손익] 미실현이익(NI)",
                    }
                )
                # Line 2: PL Entry (-)
                all_cf_entries.append(
                    {
                        "조정유형": caje_type,
                        "회사명": corp_name,
                        "계정코드": pl_acc_code,
                        "조정금액": -total_pl_impact,
                        "설명": "[비현금손익] 미실현이익(손익)",
                    }
                )
            elif caje_type == "CAJE04":
                if ni_code is None:
                    st.warning(
                        f"[{sheet_name}] CF조정 건너뜀: 당기순이익 계정 코드를 CoA에서 찾을 수 없습니다."
                    )
                    continue
                df_with_fs = df.merge(
                    coa_df_internal[["계정코드", "FS_Element"]],
                    on="계정코드",
                    how="left",
                )
                pl_rows = df_with_fs[df_with_fs["FS_Element"].isin(["X", "R"])].copy()
                bs_rows = df_with_fs[
                    df_with_fs["FS_Element"].isin(["A", "L", "E"])
                ].copy()
                if pl_rows.empty or bs_rows.empty:
                    st.warning(
                        f"[{sheet_name}] CF조정 건너뜀: 시트에서 손익(R/X) 또는 재무상태(A/L/E) 계정을 찾을 수 없습니다."
                    )
                    continue

                pl_pivot = pl_rows.pivot_table(
                    index=["계정코드", "FS_Element"],
                    columns="당기전기",
                    values="금액",
                    aggfunc="sum",
                ).fillna(0)
                if "당기" not in pl_pivot.columns:
                    pl_pivot["당기"] = 0
                if "전기" not in pl_pivot.columns:
                    pl_pivot["전기"] = 0
                pl_pivot["change"] = pl_pivot["당기"] + pl_pivot["전기"]
                pl_pivot["impact"] = pl_pivot.apply(
                    lambda r: r["change"] if r.name[1] == "X" else -r["change"], axis=1
                )
                total_pl_impact = pl_pivot["impact"].sum()

                pl_acc_code = pl_rows.iloc[0]["계정코드"]
                corp_name = df.iloc[0]["회사명"]

                # Line 1: NI Entry (+)
                all_cf_entries.append(
                    {
                        "조정유형": caje_type,
                        "회사명": corp_name,
                        "계정코드": ni_code,
                        "조정금액": total_pl_impact,
                        "설명": "[손익/재무활동] 미실현이익(NI)",
                    }
                )
                # Line 2: RE/PL Entry (-)
                all_cf_entries.append(
                    {
                        "조정유형": caje_type,
                        "회사명": corp_name,
                        "계정코드": pl_acc_code,
                        "조정금액": total_pl_impact,
                        "설명": "[손익/재무활동] 미실현이익(손익)",
                    }
                )
            elif caje_type == "CAJE97":
                if ni_code is None:
                    st.warning(
                        f"[{sheet_name}] CF조정 건너뜀: 당기순이익 계정 코드를 CoA에서 찾을 수 없습니다."
                    )
                    continue
                df_with_fs = df.merge(
                    coa_df_internal[["계정코드", "FS_Element"]],
                    on="계정코드",
                    how="left",
                )
                pl_rows = df_with_fs[df_with_fs["FS_Element"].isin(["X", "R"])].copy()
                bs_rows = df_with_fs[
                    df_with_fs["FS_Element"].isin(["A", "L", "E"])
                ].copy()
                if pl_rows.empty or bs_rows.empty:
                    st.warning(
                        f"[{sheet_name}] CF조정 건너뜀: 시트에서 손익(R/X) 또는 재무상태(A/L/E) 계정을 찾을 수 없습니다."
                    )
                    continue

                pl_pivot = pl_rows.pivot_table(
                    index=["계정코드", "FS_Element"],
                    columns="당기전기",
                    values="금액",
                    aggfunc="sum",
                ).fillna(0)
                if "당기" not in pl_pivot.columns:
                    pl_pivot["당기"] = 0
                if "전기" not in pl_pivot.columns:
                    pl_pivot["전기"] = 0
                pl_pivot["change"] = pl_pivot["당기"] + pl_pivot["전기"]
                pl_pivot["impact"] = pl_pivot.apply(
                    lambda r: r["change"] if r.name[1] == "X" else -r["change"], axis=1
                )
                total_pl_impact = pl_pivot["impact"].sum()

                pl_acc_code = pl_rows.iloc[0]["계정코드"]
                corp_name = df.iloc[0]["회사명"]

                # Line 1: NI Entry (+)
                all_cf_entries.append(
                    {
                        "조정유형": caje_type,
                        "회사명": corp_name,
                        "계정코드": ni_code,
                        "조정금액": total_pl_impact,
                        "설명": "[비현금손익] 법인세 당기손익 효과(NI)",
                    }
                )
                # Line 2: PL Entry (-)
                all_cf_entries.append(
                    {
                        "조정유형": caje_type,
                        "회사명": corp_name,
                        "계정코드": pl_acc_code,
                        "조정금액": -total_pl_impact,
                        "설명": "[비현금손익] 법인세바용",
                    }
                )
            else:
                grouped = df.groupby(["회사명", "계정코드", "설명"])
                for (corp, acc_code, desc), group in grouped:
                    pivot_df = group.pivot_table(
                        columns="당기전기", values="금액", aggfunc="sum"
                    )
                    current_amt = (
                        pivot_df["당기"].item() if "당기" in pivot_df.columns else 0
                    )
                    prior_amt = (
                        pivot_df["전기"].item() if "전기" in pivot_df.columns else 0
                    )
                    fs_element = fs_map.get(acc_code, "")

                    cf_adj_amt, cf_desc = 0, desc
                    if caje_type == "CAJE01":
                        change_amt = current_amt - prior_amt
                        if fs_element == "L":
                            cf_adj_amt = change_amt
                        else:  # For 'A' and others
                            cf_adj_amt = -change_amt
                        cf_desc = f"[운전자본] {desc}"
                    elif caje_type == "CAJE05":
                        cf_adj_amt, cf_desc = current_amt, f"[비현금손익] {desc}"

                    if abs(cf_adj_amt) > 1e-6:
                        all_cf_entries.append(
                            {
                                "조정유형": caje_type,
                                "회사명": corp,
                                "계정코드": acc_code,
                                "조정금액": cf_adj_amt,
                                "설명": cf_desc,
                            }
                        )

        bspl_cols = ["조정유형", "회사명", "계정코드", "금액", "설명", "당기전기", "FS_Element"]
        cf_cols = ["조정유형", "회사명", "계정코드", "조정금액", "설명"]

        caje_bspl_df = (
            pd.DataFrame(all_bspl_entries, columns=bspl_cols)
            if all_bspl_entries
            else pd.DataFrame(columns=bspl_cols)
        )
        caje_cf_df = (
            pd.DataFrame(all_cf_entries, columns=cf_cols)
            if all_cf_entries
            else pd.DataFrame(columns=cf_cols)
        )

        return caje_bspl_df, caje_cf_df

    if st.button(
        "🚀 최종 연결조정분개 생성 실행",
        disabled=not (
            st.session_state.adj_workflow["final_file"]
            and st.session_state.files["coa"]
        ),
    ):
        with st.spinner("최종 조정 분개를 생성하고 있습니다..."):
            try:
                coa_df = pd.read_excel(
                    st.session_state.files["coa"], sheet_name="CoA", dtype=str
                )
                caje_bspl_df, caje_cf_df = build_caje_from_template(
                    st.session_state.adj_workflow["final_file"], coa_df
                )
                st.session_state.results["caje_bspl_df"] = caje_bspl_df
                st.session_state.results["caje_cf_df"] = caje_cf_df
                st.session_state.caje_generated = True
                st.success("✅ 최종 조정 분개 생성이 완료되었습니다!")

                if not caje_bspl_df.empty and "금액" in caje_bspl_df.columns:
                    total_adj_sum = caje_bspl_df["금액"].sum()
                    if abs(total_adj_sum) > 1:
                        st.error(
                            f"❌ **[BS/PL CAJE 검증]** 조정분개 합계가 0이 아닙니다 (차대 불일치): {total_adj_sum:,.0f}"
                        )
                    else:
                        st.success(
                            f"✅ **[BS/PL CAJE 검증]** 조정분개 합계가 0으로 일치합니다."
                        )

            except Exception as e:
                st.error(f"최종 조정 분개 생성 중 오류가 발생했습니다: {e}")
                st.exception(e)

    if not st.session_state.files["coa"]:
        st.warning("먼저 사이드바에서 CoA 파일을 업로드해야 합니다.")

    if st.session_state.caje_generated:
        st.markdown("#### 📄 재무상태표/손익계산서 조정 분개 (BS/PL CAJE)")
        st.dataframe(st.session_state.results.get("caje_bspl_df"))
        st.markdown("#### 🌊 현금흐름표 조정 분개 (CF CAJE)")
        st.dataframe(st.session_state.results.get("caje_cf_df"))
        caje_excel_data = to_excel(
            {
                "CAJE_BSPL": st.session_state.results.get(
                    "caje_bspl_df", pd.DataFrame()
                ),
                "CAJE_CF": st.session_state.results.get("caje_cf_df", pd.DataFrame()),
            }
        )
        st.download_button(
            label="📥 생성된 조정 분개(CAJE) 다운로드 (.xlsx)",
            data=caje_excel_data,
            file_name="CAJE_generated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        st.info(
            "생성된 BS/PL CAJE 데이터는 '연결 재무제표' 탭의 '연결 조정' 데이터로 사용할 수 있습니다."
        )


with tab4:
    st.header("4. 외화 재무제표 환산")
    st.write(
        "외화로 작성된 재무제표(FS) 파일을 업로드하면, 지정된 환율에 따라 원화로 환산하고 결과를 표시합니다."
    )
    st.subheader("Step 1: 파일 업로드")
    st.info(
        "환산할 FS파일을 업로드해주세요. 파일의 첫 두 데이터 행에는 기말환율과 평균환율이 포함되어야 합니다."
    )
    fcfs_file = st.file_uploader("외화 FS 파일", type="xlsx", key="fcfs_uploader")
    st.subheader("Step 2: 환산 실행")
    if st.button("⚙️ 외화FS 환산 실행", disabled=not fcfs_file):
        with st.spinner("외화 재무제표를 환산하는 중입니다..."):
            try:
                log_stream = io.StringIO()
                with redirect_stdout(log_stream):
                    closing_rate, average_rate, df = read_rates_and_table(fcfs_file)
                    pre_summary = precheck_foreign_currency(df)
                    translated_df, totals_summary = translate_fcfs(
                        df, closing_rate, average_rate
                    )
                log_contents = log_stream.getvalue()
                st.session_state.fcfs_results["log"] = log_contents.strip().split("\n")
                rates_summary_df = pd.DataFrame(
                    {
                        "항목": ["기말환율", "평균환율"],
                        "값": [closing_rate, average_rate],
                    }
                )
                pre_summary_df = pd.DataFrame(
                    {"항목": list(pre_summary.keys()), "값": list(pre_summary.values())}
                )
                totals_summary_df = pd.DataFrame(
                    {
                        "항목": list(totals_summary.keys()),
                        "값": list(totals_summary.values()),
                    }
                )
                summary_df = pd.concat(
                    [rates_summary_df, pre_summary_df, totals_summary_df],
                    ignore_index=True,
                )
                summary_df["값"] = summary_df["값"].astype(str)
                st.session_state.fcfs_results["translated_df"] = translated_df
                st.session_state.fcfs_results["summary_df"] = summary_df
                st.success("🎉 외화 재무제표 환산이 완료되었습니다!")
            except Exception as e:
                st.error(f"환산 중 오류가 발생했습니다: {e}")
                st.exception(e)
    st.subheader("Step 3: 결과 확인 및 다운로드")
    if st.session_state.fcfs_results.get("log"):
        with st.expander("🔍 처리 로그 보기"):
            st.code("\n".join(st.session_state.fcfs_results["log"]))
    if st.session_state.fcfs_results.get("translated_df") is not None:
        st.markdown("#### 📄 환산된 재무제표")
        st.dataframe(st.session_state.fcfs_results["translated_df"])
        st.markdown("#### 📊 환산 요약")
        st.dataframe(st.session_state.fcfs_results["summary_df"])
        excel_data = to_excel(
            {
                "translated": st.session_state.fcfs_results["translated_df"],
                "summary": st.session_state.fcfs_results["summary_df"],
            }
        )
        st.download_button(
            label="📥 환산 결과 다운로드 (Excel)",
            data=excel_data,
            file_name="FCFS_translated.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# =================================================================================================
# --- 조정명세 차기이월 기능 ---
# =================================================================================================

def generate_carryover_adjustments(adj_file, coa_df, aje_code):
    """
    당기 조정명세 데이터를 기반으로 차기 이월 조정명세를 생성합니다.
    """
    adj_file.seek(0)
    xls = pd.ExcelFile(adj_file)
    input_sheets = {sheet_name: pd.read_excel(xls, sheet_name, dtype={"계정코드": str}) for sheet_name in xls.sheet_names}
    output_sheets = {}
    caje97_new_entries = []

    # --- 데이터 준비 ---
    if "Info" not in input_sheets:
        raise ValueError("'Info' 시트가 조정명세 파일에 없습니다.")
    info_df = input_sheets["Info"].copy()
    if '회사명' in info_df.columns:
        info_df = info_df.set_index('회사명')
    else: # 첫번째 열을 인덱스로 사용
        info_df = info_df.set_index(info_df.columns[0])

    # '당기세율' 사용
    info_df["세율_num"] = info_df["당기세율"].apply(parse_percent)
    tax_rates = info_df["세율_num"].to_dict()

    fs_map = dict(zip(coa_df["계정코드"].astype(str), coa_df["FS_Element"]))
    name_map = dict(zip(coa_df["계정코드"].astype(str), coa_df["계정명"]))

    # AJE 시트에서 계정코드/명칭 가져오기
    RE_CODE = aje_code.loc[aje_code["FS_Element"] == "E", "계정코드"].iloc[0]
    RE_NAME = aje_code.loc[aje_code["FS_Element"] == "E", "계정명"].iloc[0]
    DTL_CODE = aje_code.loc[aje_code["FS_Element"] == "L", "계정코드"].iloc[0]
    DTL_NAME = aje_code.loc[aje_code["FS_Element"] == "L", "계정명"].iloc[0]
    
    # CoA에서 비지배지분 관련 계정코드/명칭 가져오기
    nci_pl_row = coa_df[coa_df["FS_Element"] == "CR"]
    NCI_PL_CODE = nci_pl_row.iloc[0]["계정코드"] if not nci_pl_row.empty else "302000" # Fallback
    
    nci_equity_row = coa_df[coa_df["FS_Element"] == "CE"]
    NCI_EQUITY_CODE = nci_equity_row.iloc[0]["계정코드"] if not nci_equity_row.empty else "201100" # Fallback
    NCI_EQUITY_NAME = nci_equity_row.iloc[0]["계정명"] if not nci_equity_row.empty else "비지배지분"

    original_sheet_names = list(input_sheets.keys())

    # --- 시트별 차기 이월 로직 적용 ---
    for sheet_name in original_sheet_names:
        if not sheet_name.upper().startswith("CAJE"):
            if sheet_name not in output_sheets:
                 output_sheets[sheet_name] = input_sheets[sheet_name]
            continue

        caje_type = sheet_name.split("_")[0].upper()
        df = input_sheets[sheet_name].copy().dropna(how='all')
        if df.empty:
            output_sheets[sheet_name] = df
            continue
        
        df['금액'] = pd.to_numeric(df['금액'], errors='coerce').fillna(0)
        df['계정코드'] = df['계정코드'].astype(str).str.strip()
        df["FS_Element"] = df["계정코드"].map(fs_map)
        df_columns = df.columns.drop("FS_Element") if "FS_Element" in df.columns else df.columns

        # 1. CAJE00: 그대로 유지
        if caje_type == "CAJE00":
            output_sheets[sheet_name] = df.reindex(columns=df_columns)

        # 2. CAJE01: 전기 행 삭제, 당기 -> 전기
        elif caje_type == "CAJE01":
            new_df = df[df["당기전기"] == "당기"].copy()
            new_df["당기전기"] = "전기"
            output_sheets[sheet_name] = new_df.reindex(columns=df_columns)

        # 3. CAJE02: 기존 데이터 삭제 후 재작성
        elif caje_type == "CAJE02":
            new_entries = []
            df_columns = df.columns.drop("FS_Element") if "FS_Element" in df.columns else df.columns
            current_year_df = df[df["당기전기"] == "당기"].copy()

            if not current_year_df.empty:
                a_rows = current_year_df[current_year_df["FS_Element"] == "A"]
                x_rows = current_year_df[current_year_df["FS_Element"] == "X"]

                # A항목과 X항목이 모두 존재해야 로직 수행
                if not a_rows.empty and not x_rows.empty:
                    total_a_amount = a_rows["금액"].sum()

                    # A항목의 첫번째 회사명을 기준으로 사용
                    corp = a_rows.iloc[0]["회사명"]
                    
                    # X항목의 첫번째 계정코드를 사용
                    x_code = x_rows.iloc[0]["계정코드"]
                    x_name = x_rows.iloc[0]["계정명"]
                    
                    # 설명은 A항목의 첫번째 설명을 기반으로 생성
                    desc = a_rows.iloc[0].get("설명", "미실현이익")

                    # Main Adjustment
                    new_entries.append({'회사명': corp, '계정코드': x_code, '계정명': x_name, '당기전기': '전기', '금액': total_a_amount, '설명': f'{desc}'})
                    new_entries.append({'회사명': corp, '계정코드': RE_CODE, '계정명': RE_NAME, '당기전기': '전기', '금액': total_a_amount, '설명': f'{desc}'})
                    
                    # Tax Adjustment
                    tax_rate = tax_rates.get(corp, 0.0)
                    tax_effect = total_a_amount * tax_rate
                    if abs(tax_effect) > 1:
                        caje97_new_entries.append({'회사명': corp, '계정코드': RE_CODE, '계정명': RE_NAME, '당기전기': '전기', '금액': -tax_effect, '설명': f'전기 미실현이익 법인세효과 ({desc})'})
                        caje97_new_entries.append({'회사명': corp, '계정코드': DTL_CODE, '계정명': DTL_NAME, '당기전기': '전기', '금액': -tax_effect, '설명': f'전기 미실현이익 법인세효과 ({desc})'})

            output_sheets[sheet_name] = pd.DataFrame(new_entries).reindex(columns=df_columns)

        # 4. CAJE03: 모든 행 유지, X계정 -> 이익잉여금 대체
        elif caje_type == "CAJE03":
            new_df = df.copy()
            # Main Adjustment
            is_x = new_df["FS_Element"] == "X"
            new_df.loc[is_x, "금액"] = -new_df.loc[is_x, "금액"]
            new_df.loc[is_x, "계정코드"] = RE_CODE
            new_df.loc[is_x, "계정명"] = RE_NAME
            new_df.loc[is_x, "FS_Element"] = "E"
            new_df["당기전기"] = "전기"
            output_sheets[sheet_name] = new_df.reindex(columns=df_columns)
            # Tax Adjustment
            tax_effect_rows = new_df[new_df["FS_Element"].isin(["E"])].copy()
            asset_rows = new_df[new_df["FS_Element"].isin(["A"])].copy()
            if asset_rows.empty:
                continue
            asset_corp = asset_rows.iloc[0]["회사명"]
            for _, row in tax_effect_rows.iterrows():
                amount, desc = row["금액"], row.get('설명', '')
                tax_rate = tax_rates.get(asset_corp, 0.0)
                tax_effect = amount * tax_rate
                if abs(tax_effect) > 1:
                    caje97_new_entries.append({'회사명': asset_corp, '계정코드': RE_CODE, '계정명': RE_NAME, '당기전기': '전기', '금액': tax_effect, '설명': f'법인세효과 ({desc})'})
                    caje97_new_entries.append({'회사명': asset_corp, '계정코드': DTL_CODE, '계정명': DTL_NAME, '당기전기': '전기', '금액': -tax_effect, '설명': f'법인세효과 ({desc})'})

        # 5. CAJE04: 비지배지분 이월
        elif caje_type == "CAJE04":
            new_entries = []
            new_df = df.copy()
            ce_rows = new_df[new_df["FS_Element"] == "CE"]
            total_ce_amount = ce_rows["금액"].sum()
            for desc, group in ce_rows.groupby('계정코드'):
                corp = group['회사명'].iloc[0]
                ce_row = ce_rows[ce_rows['설명'] == desc]
                ce_code = ce_rows.iloc[0]["계정코드"]
                ce_name = ce_rows.iloc[0]["계정명"]
                # Main Adjustment
                new_entries.append({'회사명': corp, '계정코드': ce_code, '계정명': ce_name, '당기전기': '전기', '금액': total_ce_amount, '설명': f'배당금 조정 ({desc})'})
                new_entries.append({'회사명': corp, '계정코드': RE_CODE, '계정명': RE_NAME, '당기전기': '전기', '금액': -total_ce_amount, '설명': f'배당금 조정 ({desc})'})
            output_sheets[sheet_name] = pd.DataFrame(new_entries).reindex(columns=df_columns)

        # 6. CAJE96: X계정 -> 이익잉여금 대체
        elif caje_type == "CAJE96":
            new_df = df.copy()
            is_x = new_df["FS_Element"] == "X"
            new_df.loc[is_x, "금액"] = -new_df.loc[is_x, "금액"]
            new_df.loc[is_x, "계정코드"] = RE_CODE
            new_df.loc[is_x, "계정명"] = RE_NAME
            new_df.loc[is_x, "FS_Element"] = "E"
            new_df["당기전기"] = "전기"
            output_sheets[sheet_name] = new_df.reindex(columns=df_columns)
            # Tax Adjustment
            tax_effect_rows = new_df[new_df["FS_Element"].isin(["E"])].copy()
            for _, row in tax_effect_rows.iterrows():
                corp, amount, desc = row["회사명"], row["금액"], row.get('설명', '')
                tax_rate = tax_rates.get(corp, 0.0)
                tax_effect = amount * tax_rate
                if abs(tax_effect) > 1:
                    caje97_new_entries.append({'회사명': corp, '계정코드': RE_CODE, '계정명': RE_NAME, '당기전기': '전기', '금액': tax_effect, '설명': f'법인세효과 ({desc})'})
                    caje97_new_entries.append({'회사명': corp, '계정코드': DTL_CODE, '계정명': DTL_NAME, '당기전기': '전기', '금액': -tax_effect, '설명': f'법인세효과 ({desc})'})

        # 8. CAJE98: 당기->전기, 비지배순이익->비지배지분
        elif caje_type == "CAJE98":
            new_df = df.copy()
            is_nci_pl = new_df["계정코드"] == NCI_PL_CODE
            new_df.loc[is_nci_pl, "계정코드"] = NCI_EQUITY_CODE
            new_df.loc[is_nci_pl, "계정명"] = NCI_EQUITY_NAME
            new_df["당기전기"] = "전기"
            output_sheets[sheet_name] = new_df.reindex(columns=df_columns)
            
        # 7. CAJE97: '취득일' 등 유지, 신규 법인세 효과 추가
        elif caje_type == "CAJE97":
            preserved_rows = df[~df["당기전기"].isin(["당기", "전기"])].copy()
            output_sheets[sheet_name] = preserved_rows # 임시 저장, 나중에 신규 분개와 합침

        # 기타 조정 (CAJE05, CAJE99 등)
        else:
            output_sheets[sheet_name] = df.reindex(columns=df_columns)


    # --- CAJE97 최종 처리 ---
    def find_sheet_name_by_prefix(prefix):
        for name in original_sheet_names:
            if name.upper().startswith(prefix):
                return name
        return None # 찾지 못한 경우

    caje97_sheet_name = find_sheet_name_by_prefix("CAJE97")
    if caje97_sheet_name:
        caje97_df = output_sheets.get(caje97_sheet_name, pd.DataFrame())
        caje97_cols = input_sheets.get(caje97_sheet_name, pd.DataFrame()).columns
        if caje97_cols.empty:
            caje97_cols = ['회사명', '계정코드', '계정명', '당기전기', '금액', '설명']

        new_tax_df = pd.DataFrame(caje97_new_entries)
        final_caje97_df = pd.concat([caje97_df, new_tax_df], ignore_index=True).reindex(columns=caje97_cols)
        output_sheets[caje97_sheet_name] = final_caje97_df
    elif caje97_new_entries: # 시트는 없지만 새로 추가할 항목이 있는 경우
        caje97_sheet_name = "CAJE97_법인세조정"
        caje97_cols = ['회사명', '계정코드', '계정명', '당기전기', '금액', '설명']
        output_sheets[caje97_sheet_name] = pd.DataFrame(caje97_new_entries).reindex(columns=caje97_cols)


    # 누락된 시트가 없도록 원본 시트명 순서대로 정렬하여 반환
    final_ordered_sheets = {name: output_sheets.get(name, pd.DataFrame(columns=input_sheets.get(name, pd.DataFrame()).columns)) for name in original_sheet_names}

    return to_excel(final_ordered_sheets)

with tab3:
    st.markdown("---" )
    st.subheader("Step 6: 조정명세 차기이월 생성")
    st.write(
        "당기 조정명세 파일을 업로드하면, 차기에 반영될 전기누적 조정명세를 생성합니다."
    )

    if 'adj_workflow' not in st.session_state:
        st.session_state.adj_workflow = {}

    carryover_adj_file = st.file_uploader(
        "차기이월 할 조정명세 파일을 업로드하세요.",
        type="xlsx",
        key="carryover_uploader"
    )

    if st.button("🚀 차기이월 조정명세 생성 실행", key="run_carryover"):
        if not carryover_adj_file:
            st.warning("차기이월 할 조정명세 파일을 먼저 업로드해주세요.")
        elif not st.session_state.files["coa"]:
            st.warning("사이드바에서 CoA 파일을 먼저 업로드해주세요.")
        else:
            with st.spinner("차기이월 데이터를 생성하고 있습니다..."):
                try:
                    coa_df = pd.read_excel(st.session_state.files["coa"], sheet_name="CoA", dtype=str)
                    aje_code = pd.read_excel(st.session_state.files["coa"], sheet_name="AJE", dtype=str)

                    carryover_excel_data = generate_carryover_adjustments(
                        carryover_adj_file, coa_df, aje_code
                    )

                    st.session_state.adj_workflow["carryover_file"] = carryover_excel_data
                    st.success("🎉 차기이월 조정명세 생성이 완료되었습니다!")

                except Exception as e:
                    st.error(f"차기이월 조정명세 생성 중 오류가 발생했습니다: {e}")
                    st.exception(e)

    if "carryover_file" in st.session_state.adj_workflow and st.session_state.adj_workflow.get("carryover_file"):
        st.download_button(
            label="📥 차기이월 조정명세 다운로드 (.xlsx)",
            data=st.session_state.adj_workflow["carryover_file"],
            file_name="조정명세_입력템플릿_carryover.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
